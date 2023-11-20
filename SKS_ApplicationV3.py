

import itertools
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime
from functools import reduce
from numpy import inf
import glob
import os.path
import xlsxwriter
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
pd.options.mode.chained_assignment = None
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import win32com.client as win32
import pywintypes as py
from pywintypes import com_error
import traceback


root= tk.Tk()
root.minsize(300,300)
root.title("Generate Report")



def enter_year():
    global year_data
    #getting year entry from user
    try:
        year_data= int(year.get())
        year.delete(0, 4)
        
        return year_data
        
    except:
        tk.messagebox.showerror("Error", "Please enter a valid year", icon= "error")
        year.delete(0, 4)
        return
    
         



def enter_week():
    global w1
    #getting year entry from user
    #making customer enter the week till which it needs the report
    try:
        w1= int(Week.get())
        Week.delete(0, 4)
        
        
    
    except:
        tk.messagebox.showerror("Error", "Please enter a valid week number", icon= "error")
        Week.delete(0, 4)
        return
    

def path2():
    global path1
    
    try:
        path1 =tk.filedialog.askdirectory()
        
        
        
        if path1=="":
            tk.messagebox.showerror("Error", "Please select folder where files are located", icon='error')
        
        else:
            tk.messagebox.showinfo("Info", "Your selection:"+path1)
            
        return path1
            
    except:
        return
       


def path_finder():
    global path
    try:
        path =tk.filedialog.askdirectory()
        
        
        
        if path=="":
            tk.messagebox.showerror("Error", "Please select folder where the final file needs to be located", icon='error')
        
        else:
            tk.messagebox.showinfo("Info", "Your selection:"+path)
            
        return path
            
    except:
        return
'''
def refresh_data():
    try:
        xlapp = win32.DispatchEx("Excel.Application")
        wb = xlapp.Workbooks.Open(path1+'/tracker.xlsx')
        wb.RefreshAll()
        xlapp.CalculateUntilAsyncQueriesDone()
        wb.Save()
        xlapp.Quit()
        #tk.messagebox.showinfo("Info", "Refreshed data.", icon="info")
    
    except:
        tk.messagebox.showerror("Error", "Error refreshing data", icon="error")
'''  
def importing_data():
    global sks_data, tracker, target, sks_data_f, daily_count
    try:
        
        #importing target excel file
        
        target= pd.read_excel(path1+'/target.xlsx')
        
        #taking SKS target per hour and week
        
        target= target[["Week", "Skeptical Target/hr"]]
        
         #importing QA file
        
        
            
        
        #importing sks_data from output excel file
        
        
        file_type = '.xlsx'
        files = path1 +'/output/' +str(w1)+ file_type
        
        sks_data = pd.read_excel(files, sheet_name=0)
        
        sks_data=pd.DataFrame(sks_data)
        
        sks_data= sks_data.rename(columns={'user_id':'Login'})
        
        sks_data.drop_duplicates(subset=['asin', "Login"], inplace=True)
        
        sks_data.loc[:, "Date"]= pd.to_datetime(sks_data.loc[:,"ist"]).dt.date
        
        
        sks_data_f=sks_data.groupby(['Login', 'Date'])[['asin']].count()
        
        sks_data_f= sks_data_f.reset_index()
        
        #for daily counts
        daily_count= sks_data_f.copy()
        daily_count= daily_count.rename(columns={'asin': "Daily Count"})
        
        daily_count.sort_values('Date', ascending=True, inplace=True)
        
        sks_data_f.loc[:, "Week"]=  pd.to_datetime(sks_data_f.loc[:,"Date"]).dt.week
        
        sks_data_f.loc[:, "Week"]=sks_data_f.loc[:, "Week"]+1
        
        sks_data_f=sks_data_f.groupby(['Week', 'Login'])[['asin']].sum()
        
        sks_data_f=sks_data_f.reset_index()

        
        
        
        
        #tk.messagebox.showinfo("Info", "Successfully imported data", icon="info")
    except FileNotFoundError:
        tk.messagebox.showerror("Error", "Either output or target not found", icon='error')
    except:
        tk.messagebox.showerror("Error", "Error in importing files.", icon="error")
        
            
        



def generate_report():
    try:
        
        if path1=="":
            tk.messagebox.showerror("Error", "Please select folder where the file are located", icon='error')
        
        else:
            print(path1)
            
        

        if path=="":
            tk.messagebox.showerror("Error", "Please select folder where the final file needs to be located", icon='error')
        
        else:
            print(path)
            
        tk.messagebox.showinfo("Info", "Please wait while report is getting generated.", icon="info")
        
    
        #importing login id QA for general data
        #pd.set_option(mode.chained)
        
        files = glob.glob(path1 + '/Login.xlsx')
        max_file = max(files, key=os.path.getctime)
        
        d= pd.read_excel(max_file, sheet_name=0)
        
        d=pd.DataFrame(d)
        
        
                
        #adding week column to the original data
        d['Week']=w1
        
        #ntering year for all columns
        
        d['Year']= year_data
        
        #tracker file
        
        tracker= pd.read_excel(path1+'/Activity Tracker.xlsx', sheet_name=0)

        tracker= pd.DataFrame(tracker)
        
        tracker.fillna(0, inplace=True)
        
        tracker= tracker.rename(columns={'Emp Name': 'Name',
                                         'Total NPT':'NPT Hours',
                                         'Production Hours':'Skeptical Hours',
                                         'Leave':'Leave Hours'})
        
        df_final=tracker.copy()
        
        df_final= df_final.groupby(['Week', 'Login'])[['Skeptical Hours', 
                                                       'NPT Hours', 'Leave Hours']].sum()
        df_final=df_final.reset_index()
        
        #getting relevant columns out of leave split data

        data=df_final.copy()

        data= data[['Login', 'Week','Skeptical Hours', 'NPT Hours', 'Leave Hours']]
        
        data1= pd.merge(data,sks_data_f, how='left', on=['Login', 'Week'])
        
        data1= data1.groupby(['Week','Login'])[["asin",'Skeptical Hours', 'NPT Hours', 'Leave Hours']].sum()
        
        data1= data1.reset_index()
        
        data2= pd.merge(data1, target, how='left', on=['Week'])
        
        
        data2.loc[:, "Productive Hours"]= data2.loc[:, "Skeptical Hours"]
        
        
        data2.loc[:, "Leave Hours"]= data2.loc[:, "Leave Hours"]
        
        
        data2.loc[:, "Target"]= data2.loc[:, "Productive Hours"]*data2.loc[:, "Skeptical Target/hr"]
        
        data2.loc[:, "Achieved"]=data2.loc[:, "asin"]
        
        
        data2.loc[:, "Productivity %"]= (data2.loc[:, "Achieved"]/data2.loc[:,"Target"])*100
        
        ytd= data2.copy()
        
        ytd= pd.merge(d, ytd, on=['Login', "Week"], how='left')
        
        ytd= ytd[['Year', "Week","Name", "Login","Productive Hours", "NPT Hours", "Leave Hours",
                  "Target", "Achieved", "Productivity %"]]
        
        
        #Quality Report
        
        file_type = '.xlsx'
        files = path1 +'/QA/' +str(w1)+ file_type
        qa_file = pd.read_excel(files, sheet_name=0)
    
        qa_file=pd.DataFrame(qa_file)
        
        qa_file= qa_file.rename(columns={'user_id':'Login'})
        
        qa_file.drop_duplicates(subset=['asin', "Login"], inplace=True)
        
        
        qa_file.loc[:, "Date"]= pd.to_datetime(qa_file.loc[:,"ist"]).dt.date
        
        qa_file.loc[:, "categorized_in_correct_signal"]=['Y' if x=='BOD' else y for x,y in zip(qa_file.loc[:, "Decision"],qa_file.loc[:, "categorized_in_correct_signal"])]

        qa_file.loc[:, "captured_all_signals"]=['Y' if x=='BOD' else y for x, y in zip(qa_file.loc[:, "Decision"], qa_file.loc[:, "captured_all_signals"])]

        
        qa_data= qa_file.copy()
        
        qa_data=qa_data.groupby(['Login', 'Date'])[['asin']].count()
        
        qa_data= qa_data.reset_index()
        
        qa_data.loc[:, "Week"]= pd.to_datetime(qa_data.loc[:, 'Date']).dt.week
        
        qa_data.loc[:, "Week"]=qa_data.loc[:, "Week"]+1
        
        qa_data= qa_data.groupby(['Week', "Login"])["asin"].sum()
        
        qa_data=qa_data.reset_index()
        
        qa_data= qa_data.rename(columns={'asin': "QA Sample"})
        
        
        
        #getting fp from qa_file
        qa_file.loc[:, "FP"]= [1 if (x=='N' and y=='Y') or (x=='N' and y=='N') else 0 for x,y in zip(qa_file.loc[:, "categorized_in_correct_signal"],
                                                                                                     qa_file.loc[:, "captured_all_signals"])]
        
        qa_file.loc[:, "FN"]= [1 if x=='Y' and y=='N' else 0 for x,y in zip(qa_file.loc[:, "categorized_in_correct_signal"],
                                                                            qa_file.loc[:, "captured_all_signals"])]
        
        
        fp= qa_file.groupby(['Login'])["FP"].sum()
        fp= fp.reset_index()
        
        
        
        #getting fn from qa_file
        fn=qa_file.groupby(['Login'])["FN"].sum()
        fn= fn.reset_index()

        
        qa_data= pd.merge(qa_data, fp, how='left', on=['Login'])
        
        qa_data= pd.merge(qa_data, fn, how='left', on=['Login'])
        
        qa_data.loc[:, "FP %"]= ((qa_data.loc[:, "QA Sample"]-qa_data.loc[:, "FP"])/qa_data.loc[:, "QA Sample"])*100
        
        qa_data.loc[:, "FN %"]= ((qa_data.loc[:, "QA Sample"]-(qa_data.loc[:, "FP"]+qa_data.loc[:, "FN"]))/(qa_data.loc[:, "QA Sample"]-qa_data.loc[:, "FP"]))*100
        
        qa_data.loc[:, "Cumulative %"]= (qa_data.loc[:, "FP %"]*0.4)+(qa_data.loc[:, "FN %"]*0.6)
                                 
        #ytd
        
        final_report= pd.merge(ytd, qa_data, on=["Login", "Week"], how='left')
        
        final_report.loc[:, "Program"]= 'Skeptical Searcher'
        
        final_report= final_report.round({"Target":0, "Productivity %":2, "FP %":2, "FN %":2, "Cumulative %": 2})
        
        final_report= final_report[['Year', 'Week', 'Name', 'Login','Program', 'Productive Hours', 'NPT Hours',
                                    'Leave Hours', 'Target','Achieved', 'Productivity %', 'QA Sample',
                                    'FP', 'FP %', 'FN', 'FN %', 'Cumulative %']]
        
        #final_report.to_excel("Master_data.xlsx")
        
        #calling out master data to append current week's data
        
        
        def append_df_to_excel(filename, final_report, sheet_name, startrow=None,
                               truncate_sheet= False, 
                               **to_excel_kwargs):
            
        
            
            # Excel file doesn't exist - saving and exiting
            if not os.path.isfile(filename):
                final_report.to_excel(
                    filename,
                    sheet_name=sheet_name, 
                    startrow=startrow if startrow is not None else 0, 
                    header=True,
                    index=False,
                    **to_excel_kwargs)
                return
            
            # ignore [engine] parameter if it was passed
            if 'engine' in to_excel_kwargs:
                to_excel_kwargs.pop('engine')
        
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
        
            # try to open an existing workbook
            writer.book = load_workbook(filename)
            
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
        
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            
            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        
            if startrow is None:
                startrow = 0
        
            # write out the new sheet
            final_report.to_excel(writer, sheet_name, startrow=startrow,header=False,index=False, **to_excel_kwargs)
        
            # save the workbook
            writer.save()
        
        append_df_to_excel(path1+'/Master.xlsx',final_report, 'Master_data')
        
        files = glob.glob(path1 + '/Master.xlsx')
        max_file = max(files, key=os.path.getctime)
        
        master_data= pd.read_excel(max_file, sheet_name=0)
        
        master_data=pd.DataFrame(master_data)
        
        master_data= master_data.round({"Productive Hours":2, "NPT Hours":2,
                                        "Leave Hours":2, "Target":2, "Achieved":2,
                                        "Productivity %":2,"FP %":2, 'FN %':2,
                                        "Cumulative %":2})
        
        #T4W
        t4w= master_data.copy()
        t4w= t4w[['Year', 'Week', 'Name', 'Login', 'Productive Hours', 'NPT Hours', 'Leave Hours', 'Target', 'Achieved', 'QA Sample', 'FP', 'FN']]
        
        t4w.loc[:, "Week_new"]=""
        
        t4w = t4w.drop(t4w[(t4w.Week != w1) & (t4w.Week != w1-1) & (t4w.Week != w1-2) & (t4w.Week!=w1-3)].index)
        
        week_list= t4w['Week'].to_list()

        if w1-3 not in week_list and w1-2 in week_list and w1-1 in week_list and w1 in week_list:
            
            t4w.loc[:, "Week_new"]= str(w1-2)+'-'+str(w1)
            
            print("only 3 weeks' data is there for T4W report")
            
            tk.messagebox.showinfo("Info", "only 3 weeks data is there for T4W report.", icon='info')
            
        elif w1-3 not in week_list and w1-2 not in week_list and w1-1 in week_list and w1 in week_list:
            
            t4w.loc[:, "Week_new"]= str(w1-1)+'-'+str(w1)
            
            print("only 2 weeks' data is there for T4W report")
            
            tk.messagebox.showinfo("Info", "only 2 weeks data is there for T4W report.", icon='info')
            
        elif w1-3 not in week_list and w1-2 not in week_list and w1-1 not in week_list and w1 in week_list:
            
            t4w.loc[:, "Week_new"]= str(w1)
            print("only 1 week data is there for T4W report")
            
            tk.messagebox.showinfo("Info", "Only 1 week data is there for T4W report.", icon='info')
            
        else:
            print("All weeks data present")
            t4w.loc[:, "Week_new"]= str(w1-3)+'-'+str(w1)
            
        #print(t4w)
        print(t4w.info())
        
        t4w.loc[:,'Productive Hours'] = pd.to_numeric(t4w.loc[:,'Productive Hours'], errors='coerce')
        t4w.loc[:,'NPT Hours'] = pd.to_numeric(t4w.loc[:,'NPT Hours'], errors='coerce')
        t4w.loc[:,'Target'] = pd.to_numeric(t4w.loc[:,'Target'], errors='coerce')
        t4w.loc[:,'Achieved'] = pd.to_numeric(t4w.loc[:,'Achieved'], errors='coerce')
        t4w.loc[:,'QA Sample'] = pd.to_numeric(t4w.loc[:,'QA Sample'], errors='coerce')
        t4w.loc[:,'FP'] = pd.to_numeric(t4w.loc[:,'FP'], errors='coerce')
        t4w.loc[:,'FN'] = pd.to_numeric(t4w.loc[:,'FN'], errors='coerce')
        
        t4w=t4w.groupby(['Week_new', 'Login', 'Name','Year'])[['Productive Hours', 'NPT Hours', 'Leave Hours','Target', 'Achieved', 'QA Sample', 'FP', 'FN']].sum()
        #print(t4w)
        t4w.reset_index(inplace=True)
        
        print(t4w.info())
        
        t4w.loc[:, "Productivity %"]= (t4w.loc[:, "Achieved"]/t4w.loc[:, "Target"])*100
        
        t4w.loc[:, "FP %"]= ((t4w.loc[:, "QA Sample"]-t4w.loc[:, "FP"])/t4w.loc[:, "QA Sample"])*100
        
        t4w.loc[:, "FN %"]= ((t4w.loc[:, "QA Sample"]-(t4w.loc[:, "FP"]+t4w.loc[:, "FN"]))/(t4w.loc[:, "QA Sample"]-t4w.loc[:, "FP"]))*100
        
        t4w.loc[:, "Cumulative %"]= (t4w.loc[:, "FP %"]*0.4)+(t4w.loc[:, "FN %"]*0.6)
        
        
        
        t4w= t4w.rename(columns={'Week_new': 'Week'})
        
        t4w.loc[:, "Program"]= 'Skeptical Searcher'
        
        t4w= t4w.round({"Target":0, "Productivity %":2, "FP %":2, "FN %":2, "Cumulative %": 2})
        
        t4w= t4w[['Year', 'Week', 'Name', 'Login','Program', 'Productive Hours', 'NPT Hours',
                  'Leave Hours', 'Target','Achieved', 'Productivity %', 'QA Sample',
                  'FP', 'FP %', 'FN', 'FN %', 'Cumulative %']]
        
        
        #T4W team average
        
        team_average= t4w.copy()
        
        team_average= team_average[['Year', 'Week', 'Login', 'Name', 'Productive Hours', 'NPT Hours',
                                    'Leave Hours', 'Target', 'Achieved', 'QA Sample', 'FP', 'FN']]
        
        team_average=team_average.groupby(['Week', 'Year'])[['Productive Hours', 'NPT Hours', 'Leave Hours',
                                                                      'Target', 'Achieved', 'QA Sample', 'FP', 'FN']].sum()
        
        team_average.reset_index(inplace=True)
        
        team_average.loc[:, "Productivity %"]= (team_average.loc[:, "Achieved"]/team_average.loc[:, "Target"])*100
        
        team_average.loc[:, "FP %"]= ((team_average.loc[:, "QA Sample"]-team_average.loc[:, "FP"])/team_average.loc[:, "QA Sample"])*100
        
        team_average.loc[:, "FN %"]= ((team_average.loc[:, "QA Sample"]-(team_average.loc[:, "FP"]+team_average.loc[:, "FN"]))/(team_average.loc[:, "QA Sample"]-team_average.loc[:, "FP"]))*100
        
        team_average.loc[:, "Cumulative %"]= (team_average.loc[:, "FP %"]*0.4)+(team_average.loc[:, "FN %"]*0.6)
        
        
        
        team_average.loc[:, "Name"]='Team'
        
        team_average.loc[:, "Program"]= 'Skeptical Searcher'
        
        team_average= team_average.round({"Target":0, "Productivity %":2, "FP %":2, "FN %":2, "Cumulative %": 2})
        
        team_average= team_average[['Year', 'Week', 'Name','Program', 'Productive Hours', 'NPT Hours',
                                    'Leave Hours', 'Target','Achieved', 'Productivity %', 'QA Sample',
                                    'FP', 'FP %', 'FN', 'FN %', 'Cumulative %']]


        master_data= master_data.round({"Productive Hours":2, "NPT Hours":2,
                                        "Leave Hours":2, "Target":2, "Achieved":2,
                                        "Productivity %":2,"FP %":2, 'FN %':2,
                                        "Cumulative %":2})
        t4w= t4w.round({"Productive Hours":2, "NPT Hours":2,
                        "Leave Hours":2, "Target":2, "Achieved":2,
                        "Productivity %":2,"FP %":2, 'FN %':2,
                        "Cumulative %":2})
        
        team_average= team_average.round({"Productive Hours":2, "NPT Hours":2,
                                          "Leave Hours":2, "Target":2, "Achieved":2,
                                          "Productivity %":2,"FP %":2, 'FN %':2,
                                          "Cumulative %":2})
        
        #t4w.loc[:, "Week"]= str(w1-3)+'-'+str(w1)
        
        
        writer1 = pd.ExcelWriter(path+'/SKS_weekly_report_'+str(w1)+'.xlsx', engine='xlsxwriter')
        
        master_data.to_excel(writer1, sheet_name='YTD', index=False)
        
        t4w.to_excel(writer1, sheet_name='T4W', index=False)
        
        team_average.to_excel(writer1, sheet_name='T4W Team Average', index=False)
                
        daily_count.to_excel(writer1, sheet_name="Daily Count", index=False)
        
        writer1.save()

        
         
                
                
                
                
             
     
    except NameError:
        for e in NameError:
            print (e)
        tk.messagebox.showerror("Error", "All fields are required")
    except FileNotFoundError:
        tk.messagebox.showerror("Error", "Login or QA file not found", icon="error")
            
    except KeyError:
        for e in KeyError:
            print(e)
    
    except:
                
        tk.messagebox.showerror("Error", "Please check your files for correct data")
        
        
    else:
        tk.messagebox.showinfo("Info", "Productivity and Quality report has been generated. Check:"+path)
        
        
def window_destroy():
    root.destroy()
    
#take count from output file (unique asins against user_id)

#npts: right to sks hours 
#QA- QA sheet (unique asins)

#FP= categorized_in_correct_signal : this column is "N"

#FN= captured_all_signals: this column is "N"



    

#labels
year_label= tk.Label(root, text= "Enter Year:")  
year_label.grid(row=2, column=0,padx=5, pady=5) 

Week_label= tk.Label(root, text= "Enter Week:")
Week_label.grid(row=3, column=0,padx=5, pady=5)   


path_label1= tk.Label(root, text= "Files location:")
path_label1.grid(row=4, column=0,padx=5, pady=5)

path_label= tk.Label(root, text= "Final Report Location:")
path_label.grid(row=5, column=0,padx=5, pady=5)  



#entry boxes
year= tk.Entry(root, width= 30)
year.grid(row=2, column= 1,padx=5, pady=5)

Week= tk.Entry(root, width= 30)
Week.grid(row=3, column= 1,padx=5, pady=5)



         


#buttons
button2= tk.Button(root, text="Select folder", command= path2)
button2.grid(row=4, column=1, padx=5, pady=5)   

button3= tk.Button(root, text="Select folder", command= path_finder)
button3.grid(row=5, column=1,padx=5, pady=5)   

button_report= tk.Button( root, text= "Submit", command= lambda : [enter_year(), enter_week()])
button_report.grid(row=6, column= 0,padx=5, pady=5)

b1 = tk.Button(root, text= "Load & refresh data files", command= lambda: [importing_data()])
b1.grid(row=6, column= 1,padx=5, pady=5)


button_exit= tk.Button(root, text="Generate Report", command= generate_report)
button_exit.grid(row= 7, column= 0,padx=5, pady=5)

button_exit= tk.Button(root, text="Quit", command= window_destroy)
button_exit.grid(row= 7, column= 1,padx=5, pady=5)


root.mainloop()


