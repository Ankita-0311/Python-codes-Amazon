import itertools
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime
from functools import reduce
from numpy import inf
import glob
import os.path
import xlsxwriter
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
pd.options.mode.chained_assignment = None
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import win32com.client as win32
import pywintypes as py
from pywintypes import com_error


path=r'C:/Users/nkcho/Desktop'

path1=r'W:/My Documents/Report'

path2=r'C:/Users/nkcho/Desktop'

w1= int(input("Week no.:"))

year_data= int(input("year:"))

#Manager_name=str(input("Manager name:"))

#Team_name= str(input("Team name:"))

#importing all the files

#importing login id QA for general data
#pd.set_option(mode.chained)

manager_emails=(path1 + "/Master_Login.xlsx")

email_details = pd.read_excel(manager_emails, sheet_name=2)


email_details=pd.DataFrame(email_details)

files =(path1 + "/Master_Login.xlsx")

d1 = pd.read_excel(files, sheet_name=0)

d1=pd.DataFrame(d1)

convert_dict = {'Login':'category','Name':'category',
                'Manager':'category','Shared Associate': 'category',
                'Program':'category'}

d1 = d1.astype(convert_dict)

d2=pd.read_excel(files, sheet_name=1)

d2=pd.DataFrame(d2)


#login_list= d1.loc[:, "Login"].to_list()

d=d1[['Login', 'Name']]

#counting rows in order to repeat values based on week number
index= d.index
a= len(index)
#creating a list from 1 to the week user wants


c=(list(range(1, w1+1)))
#storing the week numbers into the list 
week= list(itertools.chain.from_iterable(itertools.repeat(x, a) for x in c))

#converting list to dataframe column
weekd=pd.DataFrame(week, columns= ['Week'])

#repating rows with login and names the no. of times user has input week numbers
df= pd.concat([d]*w1, ignore_index=True)

#df=d.copy()

d=''
#adding week column to the original data
#df['Week']=w1

df.loc[:, "Week"]=weekd
#getting month from month_data

df= pd.merge(df, d2, on=['Week'], how='left')

#ntering year for all columns

df['Year']= year_data

#changing the dtype of columns in df dataframe

convert_dict = {'Login':'category','Name':'category','Month':'category',
                'Week':'int8','Year': 'int16'}

df = df.astype(convert_dict)



#importing all the files

#importing USN file for QA

#folder_path = r'C:\Users\nkcho\Desktop\Report\USN'
file_type = '\*xlsx'
files = glob.glob(path1+'\\USN' + file_type)
max_file = max(files, key=os.path.getctime)

usno = pd.read_excel(max_file)

usno=pd.DataFrame(usno)

#importing team average file for CCR
#folder_path = r'C:\Users\nkcho\Desktop\Report\TA'
file_type = '\*xlsx'
files = glob.glob(path1+'\\TA' + file_type)
max_file = max(files, key=os.path.getctime)

ta_ccro = pd.read_excel(max_file, sheet_name=0)

ta_ccro=pd.DataFrame(ta_ccro)

#importing team average file for USN


#folder_path = r'C:\Users\nkcho\Desktop\Report\TA'
file_type = '\*xlsx'
files = glob.glob(path1+'\\TA' + file_type)
max_file = max(files, key=os.path.getctime)

ta_usno = pd.read_excel(max_file, sheet_name=1)

ta_usno=pd.DataFrame(ta_usno)

#import MCM file


#folder_path = r'C:\Users\nkcho\Desktop\Report\MCM'
file_type = '\*xlsx'
files = glob.glob(path1+'\\MCM' + file_type)
max_file = max(files, key=os.path.getctime)

mcmo = pd.read_excel(max_file)

mcmo=pd.DataFrame(mcmo)



#import ccr file


#folder_path = r'C:\Users\nkcho\Desktop\Report\CCR'
file_type = '\*xlsx'
files = glob.glob(path1+'\\CCR' + file_type)
max_file = max(files, key=os.path.getctime)

ccro = pd.read_excel(max_file, sheet_name=0)

ccro=pd.DataFrame(ccro)

#import adhoc projects


#folder_path = r'C:\Users\nkcho\Desktop\Report\adhoc'
file_type = '\*xlsx'
files = glob.glob(path1+'\\adhoc' + file_type)
max_file = max(files, key=os.path.getctime)

adhoco = pd.read_excel(max_file)

adhoco=pd.DataFrame(adhoco)

#import EIM Escalation file


#folder_path = r'C:\Users\nkcho\Desktop\Report\EIM'
file_type = '\*xlsx'
files = glob.glob(path1+'\\EIM Esc' + file_type)
max_file = max(files, key=os.path.getctime)

eimo = pd.read_excel(max_file, sheet_name=0)

eimo=pd.DataFrame(eimo)

#import EIM FP file


#folder_path = r'C:\Users\nkcho\Desktop\Report\EIM'
file_type = '\*xlsx'
files = glob.glob(path1+'\\EIM FP' + file_type)
max_file = max(files, key=os.path.getctime)

eim_o = pd.read_excel(max_file, sheet_name=0)

eim_o=pd.DataFrame(eim_o)

#EPCR, EACP and EACP Appeals
file_type = '\*xlsx'
files = glob.glob(path1+'\\EPCR' + file_type)
max_file = max(files, key=os.path.getctime)

epcr = pd.read_excel(max_file, sheet_name=0)

epcr=pd.DataFrame(epcr)

#EPCR team average values

file_type = '\*xlsx'
files = glob.glob(path1+'\\TA' + file_type)
max_file = max(files, key=os.path.getctime)

ta_epcr = pd.read_excel(max_file, sheet_name=2)

ta_epcr=pd.DataFrame(ta_epcr)

#eacp and eacp appeals

file_type = '\*xlsx'
files = glob.glob(path1+'\\EACP' + file_type)
max_file = max(files, key=os.path.getctime)

eacp_data=pd.read_excel(max_file, sheet_name=0)
eacp_data= pd.DataFrame(eacp_data)

#DCR
file_type = '\*xlsx'
files = glob.glob(path1+'\\Defective' + file_type)
max_file = max(files, key=os.path.getctime)

dcro = pd.read_excel(max_file, sheet_name=0)

dcro=pd.DataFrame(dcro)

#Coverage
file_type = '\*xlsx'
files = glob.glob(path1+'\\Coverage' + file_type)
max_file = max(files, key=os.path.getctime)

coverageo = pd.read_excel(max_file, sheet_name=0)

coverageo=pd.DataFrame(coverageo)

#define the sorter which is Month used in order to sort the data as per month

Month_sorter= ['January', 'February', 'March', 'April', 'May', 'June', 
    'July', 'August', 'September', 'October', 'November', 'December']

#import productivity tracker report for daily count as well as weekly count

file_type = '\*xlsx'
files = glob.glob(path1+'\\Tracker' + file_type)
max_file = max(files, key=os.path.getctime)

t1 = pd.read_excel(max_file, sheet_name=3)

t1=pd.DataFrame(t1)

t2 = pd.read_excel(max_file, sheet_name=2)

t2=pd.DataFrame(t2)
#retrieving relevant columns

#t1= t1.drop(t1[(t1.loc[:, "Week"]!=w1)].index)

t1= t1.groupby(['Week', 'Login'])[['CCR EN Count', 'CCR Non EN Count', 'Brand Owner Audits Count', 
                                   'Non AVP Audits EN Count', 'Non AVP Audits Non EN Count',
                                   'MCM EN Count', 'MCM Non EN Count', 'USN EN Count', 'USN Non EN Count',
                                   'EPCR EN Count', 'EPCR Non EN Count', 'EACP Count', 'EACP Appeals Count',
                                   'CCR Total Achieved','EIM FP Total Achieved',
                                   'EIM Esc Total Achieved','MCM Total Achieved', 'USN Total Achieved',
                                   'EPCR Total Achieved', 'DCR Total Achieved',
                                   'Coverage Total Achieved']].sum()

t1=t1.reset_index()

t2=t2[['Week', 'Login', 'Total','Actual Cumputed Hours','Actual Cumputed NPT',
       'Leave Hours','Weighted Productive Sum','Weighted Productivity %']]

#t2= t2.drop(t2[(t2.loc[:, "Week"]!=w1)].index)

t2['Weighted Productivity %'] = pd.to_numeric(t2['Weighted Productivity %'], errors='coerce')
t2.fillna(0, inplace= True)

#merging all the data
tracker_data= [t1,t2]

tracker= reduce(lambda left,right: pd.merge(left,right,on=['Week', 'Login']), tracker_data)
tracker=tracker.reset_index(drop=True)

#tracker= tracker.drop(tracker[(tracker.loc[:, "Week"]!=w1)].index)

tracker.reset_index(inplace=True)

tracker=tracker.drop(columns=['index'])

t1=''
t2=''

#USN File
#changing the datatype of USN
usno= usno.rename(columns={usno.columns[5]:'Week',
                           usno.columns[2]:'Login',
                           usno.columns[10]:'USN QA %'})




#adding total errors columns
usno.loc[:,'USN Total Errors']= usno.iloc[:,7]+usno.iloc[:,8]
usno.loc[:,'USN QA Sample']= usno.iloc[:,6]

#gettng the relevant columns from USN file
usn1= usno[usno.columns[(2,5,10,11,12), ]].copy()

usn1.fillna(0, inplace=True)

usn1.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'USN QA Sample':'int16', 'USN Total Errors':'int16',
                'USN QA %':'float16','Week':'int16'}

usn1 = usn1.astype(convert_dict)



#total USN count from tracker file
tracker_usn= tracker[['Week', 'Login', 'USN Total Achieved']]

tracker_usn=tracker_usn.rename(columns={'USN Total Achieved': "USN Tagged Volume"})


#getting team average values
ta_usno= ta_usno.rename(columns={ta_usno.columns[0]:'Week'})
#ta_usno= ta_usno.drop(ta_usno[(ta_usno.loc[:, "Week"]!=w1)].index)

ta_usn1= ta_usno[ta_usno.columns[(0,7,9), ]].copy()
ta_usn1.loc[:, 'USN Team Average']= ta_usn1.iloc[:, 2]/ta_usn1.iloc[:, 1]
ta_usn1= ta_usn1[['Week', 'USN Team Average']]




#merging USN with df file
df= pd.merge(df, usn1, how='left', on=['Week', 'Login'])

#merging with tracker data
df= pd.merge(df, tracker_usn,how='left', on= ['Week', 'Login'])

#merging with team average data
df= pd.merge(df, ta_usn1, how='left', on=['Week'])


df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average']]

df['USN Tagged Volume'].fillna(0, inplace=True)
df['USN QA Sample'].fillna(0, inplace=True)
df['USN Total Errors'].fillna(0, inplace=True)
df['USN QA %'].fillna(0, inplace=True)
df['USN Team Average'].fillna(0, inplace=True)


convert_dict = {'USN Tagged Volume':'int16', 'USN QA Sample':'int16', 'USN Total Errors':'int16',
                'USN QA %':'float16','USN Team Average':'float16'}

df = df.astype(convert_dict)

#MCM file

#converting object to float in MCM file
mcmo.iloc[:,11] = pd.to_numeric(mcmo.iloc[:,11], errors='coerce')
mcmo.iloc[:,9] = pd.to_numeric(mcmo.iloc[:,9], errors='coerce')
mcmo.iloc[:,10] = pd.to_numeric(mcmo.iloc[:,10], errors='coerce')



#getting team average for MCM
team_average= mcmo.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= mcmo.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg_mcm= pd.merge(team_average, team_average1, on=['Week'])
team_avg_mcm= pd.DataFrame(team_avg_mcm)
team_avg=team_avg_mcm.copy()

team_avg.loc[:,'MCM Team Average']= (team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'MCM Team Average']]
#merging with MCM file
mcm= mcmo.copy()



#rename columns
mcm= mcm.rename(columns={mcm.columns[4]:'Login',
                         mcm.columns[3]:'Week',
                         mcm.columns[10]:'MCM Total Errors',
                         mcm.columns[11]:'MCM QA %',
                         mcm.columns[9]:'MCM QA Sample'})

#deleting irrelevant columns for mcm file

mcm= mcm[['Week','Login', 'MCM Total Errors', 'MCM QA %', 'MCM QA Sample']]

mcm.fillna(0, inplace=True)
mcm.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'MCM QA Sample':'int16', 'MCM Total Errors':'int16',
                'MCM QA %':'float16','Week':'int16'}

mcm = mcm.astype(convert_dict)

#total MCM count from tracker file
tracker_mcm= tracker[['Week', 'Login', 'MCM Total Achieved']]
tracker_mcm= tracker_mcm.rename(columns={'MCM Total Achieved':'MCM Tagged Volume'})


#merging with mcm file to the final mcm file
df= pd.merge(df, mcm, how='left', on= ['Week', 'Login'])

#merging the final mcm with team average data
df= pd.merge(df, team_avg, how='left', on=['Week'])

team_avg=""

#merging with final mcm with tracker data
df= pd.merge(df, tracker_mcm, how='left', on=['Week', 'Login'])





df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average']]

df['MCM Tagged Volume'].fillna(0, inplace=True)
df['MCM QA Sample'].fillna(0, inplace=True)
df['MCM Total Errors'].fillna(0, inplace=True)
df['MCM QA %'].fillna(0, inplace=True)
df['MCM Team Average'].fillna(0, inplace=True)


convert_dict = {'MCM Tagged Volume':'int16', 'MCM QA Sample':'int16', 'MCM Total Errors':'int16',
                'MCM QA %':'float16','MCM Team Average':'float16'}

df = df.astype(convert_dict)

#CCR file
#CCR file renaming
ccro= ccro.rename(columns={ccro.columns[4]:'Login',
                           ccro.columns[3]:'Week',
                           ccro.columns[9]: 'CCR Total Errors',
                           ccro.columns[8]:'CCR QA Sample',
                           ccro.columns[10]:'CCR QA %'})
#ccro= ccro.drop(ccro[(ccro.loc[:, "Week"]!=w1)].index)

ccr1=ccro.copy()

ccr1.fillna(0, inplace=True)
ccr1.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'CCR QA Sample':'int16', 'CCR Total Errors':'int16',
                'CCR QA %':'float16','Week':'int16'}

ccr1 = ccr1.astype(convert_dict)


#gettng the relevant columns from CCR file
ccr1= ccr1[ccr1.columns[(3,4,8,9,10), ]]



#getting relevant columns for TA
ta_ccr1= ta_ccro[ta_ccro.columns[(0,8,9), ]].copy()
ta_ccr1= ta_ccr1.rename(columns={ta_ccr1.columns[0]:'Week'})
#ta_ccr1= ta_ccr1.drop(ta_ccr1[(ta_ccr1.loc[:, "Week"]!=w1)].index)

ta_ccr1.loc[:, 'CCR Team Average']= ta_ccr1.iloc[:, 1]/ ta_ccr1.iloc[:, 2]
ta_ccr1= ta_ccr1[['Week','CCR Team Average']]

#total MCM count from tracker file
tracker_ccr= tracker[['Week', 'Login', 'CCR EN Count', 'CCR Non EN Count']]
tracker_ccr.loc[:,'CCR Tagged Volume']= tracker_ccr.iloc[:,2]+tracker_ccr.iloc[:,3]

tracker_ccr= tracker_ccr[["Week", "Login","CCR Tagged Volume"]]


#merging ccr1 file with df2

df= pd.merge(df, ccr1, how='left', on= ['Week', 'Login'])

#merging ccr final with the columns with the tracker report

df= pd.merge(df, tracker_ccr, how= 'left', on=['Week', 'Login'])

ta_ccr1= ta_ccr1.astype({'Week': int})


#merging ccr final with TA file for CCR
df= pd.merge(df, ta_ccr1, how='left', on=['Week'])




df=df[['Login', 'Name', 'Week', 'Month','Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average']]

df['CCR Tagged Volume'].fillna(0, inplace=True)
df['CCR QA Sample'].fillna(0, inplace=True)
df['CCR Total Errors'].fillna(0, inplace=True)
df['CCR QA %'].fillna(0, inplace=True)
df['CCR Team Average'].fillna(0, inplace=True)


convert_dict = {'CCR Tagged Volume':'int16', 'CCR QA Sample':'int16', 'CCR Total Errors':'int16',
                'CCR QA %':'float16','CCR Team Average':'float16'}

df = df.astype(convert_dict)



#EIM Escalations file
eimo= eimo.rename(columns={eimo.columns[3]: 'Week',
                           eimo.columns[4]: 'Login',
                           eimo.columns[12]: 'EIM Esc QA %',
                           eimo.columns[7]: 'EIM Esc QA Sample'})

#eimo= eimo.drop(eimo[(eimo.loc[:, "Week"]!=w1)].index)

eimo.loc[:, "EIM Esc Total Errors"]=eimo.iloc[:, 10]+eimo.iloc[:, 11]

#getting team average for EIM Esc
team_avg= eimo.groupby(['Week'])[['EIM Esc QA Sample','EIM Esc Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EIM Esc Team Average']= (team_avg['EIM Esc QA Sample']- team_avg['EIM Esc Total Errors'])/team_avg['EIM Esc QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'EIM Esc Team Average']]

eim1= eimo[['Week','Login', 'EIM Esc Total Errors', 'EIM Esc QA %'
            , 'EIM Esc QA Sample']].copy()
eim1.fillna(0, inplace=True)

convert_dict = {'EIM Esc QA Sample':'int16', 'EIM Esc Total Errors':'int16',
                'EIM Esc QA %':'float16','Week':'int16'}

eim1 = eim1.astype(convert_dict)


tracker_eim= tracker[['Week', 'Login', 'EIM Esc Total Achieved']]


eim1.fillna(0, inplace= True)



#merging with the df
df= pd.merge(df, eim1,how='left', on=['Week', 'Login'])

df=pd.merge(df, team_avg, how='left', on=['Week'])

df= pd.merge(df, tracker_eim,how='left', on=['Week', 'Login'])

df.reset_index(inplace=True)

df= df.rename(columns= {'EIM Esc Total Achieved':'EIM Esc Tagged Volume'})



df=df[['Login', 'Name', 'Week', 'Month','Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average']]

df['EIM Esc Tagged Volume'].fillna(0, inplace=True)
df['EIM Esc QA Sample'].fillna(0, inplace=True)
df['EIM Esc Total Errors'].fillna(0, inplace=True)
df['EIM Esc QA %'].fillna(0, inplace=True)
df['EIM Esc Team Average'].fillna(0, inplace=True)


convert_dict = {'EIM Esc Tagged Volume':'int16', 'EIM Esc QA Sample':'int16', 'EIM Esc Total Errors':'int16',
                'EIM Esc QA %':'float16','EIM Esc Team Average':'float16'}

df = df.astype(convert_dict)


#EIM FP file
eim_o= eim_o.rename(columns={eim_o.columns[3]: 'Week',
                             eim_o.columns[4]: 'Login',
                             eim_o.columns[10]: 'EIM FP Total Errors',
                             eim_o.columns[11]: 'EIM FP QA %',
                             eim_o.columns[8]: 'EIM FP QA Sample'})


#eim_o= eim_o.drop(eim_o[(eim_o.loc[:, "Week"]!=w1)].index)

#getting team average for EIM Esc
team_avg= eim_o.groupby(['Week'])[['EIM FP QA Sample','EIM FP Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EIM FP Team Average']= (team_avg['EIM FP QA Sample']- team_avg['EIM FP Total Errors'])/team_avg['EIM FP QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'EIM FP Team Average']]



eim1= eim_o[['Week','Login', 'EIM FP Total Errors', 'EIM FP QA %'
, 'EIM FP QA Sample']].copy()

eim1.fillna(0, inplace=True)

eim1.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'EIM FP QA Sample':'int16', 'EIM FP Total Errors':'int16',
                'EIM FP QA %':'float16','Week':'int16'}

eim1 = eim1.astype(convert_dict)


tracker_eim= tracker[['Week','Login','EIM FP Total Achieved' ]]

tracker_eim= tracker_eim.rename(columns= {'EIM FP Total Achieved':'EIM FP Tagged Volume'})

eim1= eim1[['Week', 'Login', 'EIM FP Total Errors',
            'EIM FP QA %', 'EIM FP QA Sample']]

eim1.fillna(0, inplace= True)

eim1.replace(to_replace= '-', value= 0, inplace= True)

#merging with the df3
df= pd.merge(df, eim1, how='left',on=['Week', 'Login'])

df= pd.merge(df, tracker_eim, how='left',on=['Week', 'Login'])

df= pd.merge(df, team_avg, how='left',on=['Week'])

df.reset_index(inplace=True)



df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average']]

df['EIM FP Tagged Volume'].fillna(0, inplace=True)
df['EIM FP QA Sample'].fillna(0, inplace=True)
df['EIM FP Total Errors'].fillna(0, inplace=True)
df['EIM FP QA %'].fillna(0, inplace=True)
df['EIM FP Team Average'].fillna(0, inplace=True)


convert_dict = {'EIM FP Tagged Volume':'int16', 'EIM FP QA Sample':'int16', 'EIM FP Total Errors':'int16',
                'EIM FP QA %':'float16','EIM FP Team Average':'float16'}

df = df.astype(convert_dict)



#adhoc file
#replacing values in adhoc
adhoco.replace(to_replace= '-', value= 0, inplace= True)
adhoco.fillna(0, inplace=True)

#converting object to float in MCM file

adhoco.iloc[:,4] = pd.to_numeric(adhoco.iloc[:,4], errors='coerce')
adhoco.iloc[:,5] = pd.to_numeric(adhoco.iloc[:,5], errors='coerce')
adhoco.iloc[:,6] = pd.to_numeric(adhoco.iloc[:,6], errors='coerce')
adhoco.iloc[:,7] = pd.to_numeric(adhoco.iloc[:,7], errors='coerce')
adhoco.iloc[:,8] = pd.to_numeric(adhoco.iloc[:,8], errors='coerce')
adhoco.iloc[:,9] = pd.to_numeric(adhoco.iloc[:,9], errors='coerce')
adhoco.iloc[:,10] = pd.to_numeric(adhoco.iloc[:,10], errors='coerce')
adhoco.iloc[:,11] = pd.to_numeric(adhoco.iloc[:,11], errors='coerce')
adhoco.iloc[:,12] = pd.to_numeric(adhoco.iloc[:,12], errors='coerce')

#renaming log in columns
adhoco= adhoco.rename(columns={adhoco.columns[3]: 'Login',
                adhoco.columns[1]: 'Week',
                adhoco.columns[4]: 'BO QA Sample Audits',
                adhoco.columns[5]: 'BO Total Errors',
                adhoco.columns[6]: 'BO QA %',
                adhoco.columns[7]:'BO Team Average',
                adhoco.columns[8]: 'Transparency Tagged Volume',
                adhoco.columns[9]: 'Transparency QA Sample Audits',
                adhoco.columns[10]: 'Transparency Total Errors',
                adhoco.columns[11]: 'Transparency QA %',
                adhoco.columns[12]:'Transparency Team Average'})


adhoc1= adhoco[['Week', 'Login','BO QA Sample Audits', 'BO Total Errors','BO QA %',
                
                'Transparency Tagged Volume', 'Transparency QA Sample Audits', 
                'Transparency Total Errors', 'Transparency QA %']]

adhoc1.fillna(0, inplace=True)

convert_dict = {'BO QA Sample Audits':'int16', 'BO Total Errors':'int16',
                'BO QA %':'float16','Week':'int16',
                
                'Transparency QA Sample Audits':'int16', 'Transparency Total Errors':'int16',
                'Transparency QA %':'float16'}


adhoc1 = adhoc1.astype(convert_dict)


adhoc2=adhoco[['Week','BO Team Average', 'Transparency Team Average']]

adhoc2=adhoc2.drop_duplicates(subset=['Week'])
adhoc2.reset_index(drop=True)
#adhoc2= adhoc2.drop(adhoc2[(adhoc2.loc[:, "Week"]!=w1)].index)

adhoc2.fillna(0, inplace=True)

#brand owner tagged volume from tracker data
tagged_volume= tracker[['Week', 'Login', 'Brand Owner Audits Count']]

tagged_volume= tagged_volume.rename(columns={'Brand Owner Audits Count': 'BO Tagged Volume'})

tagged_volume=tagged_volume[['Week', 'Login', 'BO Tagged Volume']]

adhoc=pd.merge(adhoc1,tagged_volume, how='left', on=['Week', 'Login'])

#adhoc= adhoc.drop(adhoc[(adhoc.loc[:, "Week"]!=w1)].index)

#adding bo column to the adhoc data

df= pd.merge(df, adhoc, how='left', on=['Week','Login'])

df= pd.merge(df, adhoc2, how='left', on=['Week'])

df.reset_index(inplace=True)




df=df[['Login', 'Name', 'Week', 'Month','Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average',
       'BO Tagged Volume', 'BO QA Sample Audits', 'BO Total Errors', 'BO QA %',
       'BO Team Average',
       'Transparency Tagged Volume', 'Transparency QA Sample Audits', 'Transparency Total Errors',
       'Transparency QA %', 'Transparency Team Average']]

df['BO Tagged Volume'].fillna(0, inplace=True)
df['BO QA Sample Audits'].fillna(0, inplace=True)
df['BO Total Errors'].fillna(0, inplace=True)
df['BO QA %'].fillna(0, inplace=True)
df['BO Team Average'].fillna(0, inplace=True)



df['Transparency Tagged Volume'].fillna(0, inplace=True)
df['Transparency QA Sample Audits'].fillna(0, inplace=True)
df['Transparency Total Errors'].fillna(0, inplace=True)
df['Transparency QA %'].fillna(0, inplace=True)
df['Transparency Team Average'].fillna(0, inplace=True)


convert_dict = {'BO Tagged Volume':'int16', 'BO QA Sample Audits':'int16', 'BO Total Errors':'int16',
                'BO QA %':'float16','BO Team Average':'float16',
                'Transparency Tagged Volume':'int16', 'Transparency QA Sample Audits':'int16', 'Transparency Total Errors':'int16',
                'Transparency QA %':'float16','Transparency Team Average':'float16'}

df = df.astype(convert_dict)



#DCR program

#dcro= dcro.drop(dcro[(dcro.loc[:, "Week"]!=w1)].index)

#getting team average for MCM
team_average= dcro.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= dcro.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Week'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'DCR Team Average']= (team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'DCR Team Average']]
#merging with MCM file
dcr= dcro.copy()



#rename columns
dcr= dcr.rename(columns={dcr.columns[4]:'Login',
            dcr.columns[3]:'Week',
            dcr.columns[9]:'DCR Total Errors',
            dcr.columns[11]:'DCR QA %',
            dcr.columns[8]:'DCR QA Sample'})

#deleting irrelevant columns for mcm file

dcr= dcr[['Week','Login', 'DCR Total Errors', 'DCR QA %', 'DCR QA Sample']]

dcr.fillna(0, inplace=True)

dcr.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'DCR QA Sample':'int16', 'DCR Total Errors':'int16',
                'DCR QA %':'float16','Week':'int16'}

dcr = dcr.astype(convert_dict)


#total MCM count from tracker file
tracker_dcr= tracker[['Week', 'Login', 'DCR Total Achieved']]
tracker_dcr= tracker_dcr.rename(columns={'DCR Total Achieved':'DCR Tagged Volume'})


#merging with mcm file to the final mcm file
df= pd.merge(df, dcr, how='left', on= ['Week', 'Login'])

#merging the final mcm with team average data
df= pd.merge(df, team_avg, how='left', on=['Week'])

#merging with final mcm with tracker data
df= pd.merge(df, tracker_dcr, how='left', on=['Week', 'Login'])



df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average',
       'BO Tagged Volume', 'BO QA Sample Audits', 'BO Total Errors', 'BO QA %',
       'BO Team Average',
       'Transparency Tagged Volume', 'Transparency QA Sample Audits', 'Transparency Total Errors',
       'Transparency QA %', 'Transparency Team Average',
       'DCR Tagged Volume', 'DCR QA Sample', 'DCR Total Errors',
       'DCR QA %', 'DCR Team Average']]

df['DCR Team Average'].fillna(0, inplace=True)
df['DCR QA %'].fillna(0, inplace=True)
df['DCR Tagged Volume'].fillna(0, inplace=True)
df['DCR QA Sample'].fillna(0, inplace=True)
df['DCR Total Errors'].fillna(0, inplace=True)


convert_dict = {'DCR Tagged Volume':'int16', 'DCR QA Sample':'int16', 'DCR Total Errors':'int16',
                'DCR QA %':'float16','DCR Team Average':'float16'}

df = df.astype(convert_dict)




#coverage

#coverageo= coverageo.drop(coverageo[(coverageo.loc[:, "Week"]!=w1)].index)

#getting team average for Coverage
team_average= coverageo.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= coverageo.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Week'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'Coverage Team Average']= (team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'Coverage Team Average']]

#merging with main file
coverage= coverageo.copy()



#rename columns
coverage= coverage.rename(columns={coverage.columns[4]:'Login',
                                   coverage.columns[3]:'Week',
                                   coverage.columns[10]:'Coverage Total Errors',
                                   coverage.columns[11]:'Coverage QA %',
                                   coverage.columns[9]:'Coverage QA Sample'})

#deleting irrelevant columns for coverage file

coverage= coverage[['Week','Login', 'Coverage Total Errors', 'Coverage QA %', 'Coverage QA Sample']]

coverage.fillna(0, inplace=True)

coverage.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'Coverage QA Sample':'int16', 'Coverage Total Errors':'int16',
                'Coverage QA %':'float16','Week':'int16'}

coverage = coverage.astype(convert_dict)

#total coverage count from tracker file
tracker_coverage= tracker[['Week', 'Login', 'Coverage Total Achieved']]
tracker_coverage= tracker_coverage.rename(columns={'Coverage Total Achieved':'Coverage Tagged Volume'})


#merging with mcm file to the final coverage file
df= pd.merge(df, coverage, how='left', on= ['Week', 'Login'])

#merging the final coverage with team average data
df= pd.merge(df, team_avg, how='left', on=['Week'])

#merging with final coverage with tracker data
df= pd.merge(df, tracker_coverage, how='left', on=['Week', 'Login'])




df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average',
       'BO Tagged Volume', 'BO QA Sample Audits', 'BO Total Errors', 'BO QA %',
       'BO Team Average',
       'Transparency Tagged Volume', 'Transparency QA Sample Audits', 'Transparency Total Errors',
       'Transparency QA %', 'Transparency Team Average',
       'DCR Tagged Volume', 'DCR QA Sample', 'DCR Total Errors',
       'DCR QA %', 'DCR Team Average',
       'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors',
       'Coverage QA %', 'Coverage Team Average']]

df['Coverage Team Average'].fillna(0, inplace=True)
df['Coverage QA %'].fillna(0, inplace=True)
df['Coverage Tagged Volume'].fillna(0, inplace=True)
df['Coverage QA Sample'].fillna(0, inplace=True)
df['Coverage Total Errors'].fillna(0, inplace=True)


convert_dict = {'Coverage Tagged Volume':'int16', 'Coverage QA Sample':'int16', 'Coverage Total Errors':'int16',
                'Coverage QA %':'float16','Coverage Team Average':'float16'}

df = df.astype(convert_dict)





#epcr

epcr1= epcr.copy()

epcr1= epcr1.rename(columns={epcr1.columns[1]:'Week',
                            epcr1.columns[2]:'Login',
                            epcr1.columns[4]:'EPCR QA Sample',
                            epcr1.columns[5]: 'EPCR Total Errors',
                            epcr1.columns[6]:'EPCR QA %'})

#epcr1= epcr1.drop(epcr1[(epcr1.loc[:, "Week"]!=w1)].index)

epcr1=epcr1[['Week', 'Login', 'EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %']]

epcr1.fillna(0, inplace=True)

epcr1.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'EPCR QA Sample':'int16', 'EPCR Total Errors':'int16',
                'EPCR QA %':'float16','Week':'int16'}

epcr1 = epcr1.astype(convert_dict)


#getting team average for EPCR
team_average= epcr1.groupby(['Week'])[['EPCR QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= epcr1.groupby(['Week'])[['EPCR Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Week'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EPCR Team Average']= (team_avg['EPCR QA Sample']- team_avg['EPCR Total Errors'])/team_avg['EPCR QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'EPCR Team Average']]

#total coverage count from tracker file
tracker_epcr= tracker[['Week', 'Login', 'EPCR EN Count', 'EPCR Non EN Count']]


tracker_epcr.loc[:,'EPCR Tagged Volume']= tracker_epcr.loc[:, "EPCR EN Count"]+tracker_epcr.loc[:, "EPCR Non EN Count"]


#merging with mcm file to the final coverage file
df= pd.merge(df, epcr1, how='left', on= ['Week', 'Login'])

#merging the final coverage with team average data
df= pd.merge(df, team_avg, how='left', on=['Week'])

#merging with final coverage with tracker data
df= pd.merge(df, tracker_epcr, how='left', on=['Week', 'Login'])

df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average',
       'BO Tagged Volume', 'BO QA Sample Audits', 'BO Total Errors', 'BO QA %',
       'BO Team Average',
       'Transparency Tagged Volume', 'Transparency QA Sample Audits', 'Transparency Total Errors',
       'Transparency QA %', 'Transparency Team Average',
       'DCR Tagged Volume', 'DCR QA Sample', 'DCR Total Errors',
       'DCR QA %', 'DCR Team Average',
       'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors',
       'Coverage QA %', 'Coverage Team Average',
       'EPCR Tagged Volume', 'EPCR QA Sample', 'EPCR Total Errors',
       'EPCR QA %', 'EPCR Team Average']]

df['EPCR Team Average'].fillna(0, inplace=True)
df['EPCR QA %'].fillna(0, inplace=True)
df['EPCR Tagged Volume'].fillna(0, inplace=True)
df['EPCR QA Sample'].fillna(0, inplace=True)
df['EPCR Total Errors'].fillna(0, inplace=True)


convert_dict = {'EPCR Tagged Volume':'int16', 'EPCR QA Sample':'int16', 'EPCR Total Errors':'int16',
                'EPCR QA %':'float16','EPCR Team Average':'float16'}

df = df.astype(convert_dict)





#EACP

eacp1= eacp_data.copy()

#eacp1= eacp1.drop(eacp1[(eacp1.loc[:, "Week"]!=w1)].index)

eacp1= eacp1.drop(eacp1[(eacp1.loc[:, "Program"]!='EACP')].index)

eacp1= eacp1.rename(columns={eacp1.columns[1]:'Week',
                            eacp1.columns[2]:'Login',
                            eacp1.columns[4]:'EACP QA Sample',
                            eacp1.columns[5]: 'EACP Total Errors',
                            eacp1.columns[7]:'EACP QA %'})


eacp1=eacp1[['Week', 'Login', 'EACP QA Sample', 'EACP Total Errors', 'EACP QA %']]

eacp1.fillna(0, inplace=True)
eacp1.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'EACP QA Sample':'int16', 'EACP Total Errors':'int16',
                'EACP QA %':'float16','Week':'int16'}

eacp1 = eacp1.astype(convert_dict)



#getting team average for EPCR
team_average= eacp1.groupby(['Week'])[['EACP QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp1.groupby(['Week'])[['EACP Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Week'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Team Average']= (team_avg['EACP QA Sample']- team_avg['EACP Total Errors'])/team_avg['EACP QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'EACP Team Average']]

#total coverage count from tracker file
tracker_eacp= tracker[['Week', 'Login', 'EACP Count']]


tracker_eacp.loc[:,'EACP Tagged Volume']= tracker_eacp.loc[:, "EACP Count"]


#merging with mcm file to the final coverage file
df= pd.merge(df, eacp1, how='left', on= ['Week', 'Login'])

#merging the final coverage with team average data
df= pd.merge(df, team_avg, how='left', on=['Week'])

#merging with final coverage with tracker data
df= pd.merge(df, tracker_eacp, how='left', on=['Week', 'Login'])

df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average',
       'BO Tagged Volume', 'BO QA Sample Audits', 'BO Total Errors', 'BO QA %',
       'BO Team Average',
       'Transparency Tagged Volume', 'Transparency QA Sample Audits', 'Transparency Total Errors',
       'Transparency QA %', 'Transparency Team Average',
       'DCR Tagged Volume', 'DCR QA Sample', 'DCR Total Errors',
       'DCR QA %', 'DCR Team Average',
       'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors',
       'Coverage QA %', 'Coverage Team Average',
       'EPCR Tagged Volume', 'EPCR QA Sample', 'EPCR Total Errors',
       'EPCR QA %', 'EPCR Team Average',
       'EACP Tagged Volume', 'EACP QA Sample', 'EACP Total Errors',
       'EACP QA %', 'EACP Team Average']]

df['EACP Team Average'].fillna(0, inplace=True)
df['EACP QA %'].fillna(0, inplace=True)
df['EACP Tagged Volume'].fillna(0, inplace=True)
df['EACP QA Sample'].fillna(0, inplace=True)
df['EACP Total Errors'].fillna(0, inplace=True)


convert_dict = {'EACP Tagged Volume':'int16', 'EACP QA Sample':'int16', 'EACP Total Errors':'int16',
                'EACP QA %':'float16','EACP Team Average':'float16'}

df = df.astype(convert_dict)



#EACP appeals

eacp_appeals= eacp_data.copy()

#eacp_appeals= eacp_appeals.drop(eacp_appeals[(eacp_appeals.loc[:, "Week"]!=w1)].index)

eacp_appeals= eacp_appeals.drop(eacp_appeals[(eacp_appeals.loc[:, "Program"]!='EACP Appeals')].index)

eacp_appeals= eacp_appeals.rename(columns={eacp_appeals.columns[1]:'Week',
                                           eacp_appeals.columns[2]:'Login',
                                           eacp_appeals.columns[4]:'EACP Appeals QA Sample',
                                           eacp_appeals.columns[5]: 'EACP Appeals Total Errors',
                                           eacp_appeals.columns[7]:'EACP Appeals QA %'})


eacp_appeals=eacp_appeals[['Week', 'Login', 'EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %']]

eacp_appeals.fillna(0, inplace=True)

eacp_appeals.replace(to_replace= '-', value= 0, inplace= True)

convert_dict = {'EACP Appeals QA Sample':'int16', 'EACP Appeals Total Errors':'int16',
                'EACP Appeals QA %':'float16','Week':'int16'}

eacp_appeals = eacp_appeals.astype(convert_dict)

#getting team average for EPCR
team_average= eacp_appeals.groupby(['Week'])[['EACP Appeals QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp_appeals.groupby(['Week'])[['EACP Appeals Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Week'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Appeals Team Average']= (team_avg['EACP Appeals QA Sample']- team_avg['EACP Appeals Total Errors'])/team_avg['EACP Appeals QA Sample']
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Week', 'EACP Appeals Team Average']]

#total coverage count from tracker file
tracker_eacp_appeals= tracker[['Week', 'Login', 'EACP Appeals Count']]


tracker_eacp_appeals.loc[:,'EACP Appeals Tagged Volume']= tracker_eacp_appeals.loc[:, "EACP Appeals Count"]


#merging with mcm file to the final coverage file
df= pd.merge(df, eacp_appeals, how='left', on= ['Week', 'Login'])

#merging the final coverage with team average data
df= pd.merge(df, team_avg, how='left', on=['Week'])

#merging with final coverage with tracker data
df= pd.merge(df, tracker_eacp_appeals, how='left', on=['Week', 'Login'])

df=df[['Login', 'Name', 'Week','Month', 'Year',
       'USN Tagged Volume', 'USN QA Sample', 'USN Total Errors',
       'USN QA %','USN Team Average',
       'MCM Tagged Volume', 'MCM QA Sample', 'MCM Total Errors','MCM QA %',
       'MCM Team Average',
       'CCR Tagged Volume', 'CCR QA Sample', 'CCR Total Errors','CCR QA %',
       'CCR Team Average',
       'EIM Esc Tagged Volume', 'EIM Esc QA Sample', 'EIM Esc Total Errors','EIM Esc QA %',
       'EIM Esc Team Average',
       'EIM FP Tagged Volume', 'EIM FP QA Sample', 'EIM FP Total Errors','EIM FP QA %',
       'EIM FP Team Average',
       'BO Tagged Volume', 'BO QA Sample Audits', 'BO Total Errors', 'BO QA %',
       'BO Team Average',
       'Transparency Tagged Volume', 'Transparency QA Sample Audits', 'Transparency Total Errors',
       'Transparency QA %', 'Transparency Team Average',
       'DCR Tagged Volume', 'DCR QA Sample', 'DCR Total Errors',
       'DCR QA %', 'DCR Team Average',
       'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors',
       'Coverage QA %', 'Coverage Team Average',
       'EPCR Tagged Volume', 'EPCR QA Sample', 'EPCR Total Errors',
       'EPCR QA %', 'EPCR Team Average',
       'EACP Tagged Volume', 'EACP QA Sample', 'EACP Total Errors',
       'EACP QA %', 'EACP Team Average',
       'EACP Appeals Tagged Volume', 'EACP Appeals QA Sample', 'EACP Appeals Total Errors',
       'EACP Appeals QA %', 'EACP Appeals Team Average']]

df['EACP Appeals Team Average'].fillna(0, inplace=True)
df['EACP Appeals QA %'].fillna(0, inplace=True)
df['EACP Appeals Tagged Volume'].fillna(0, inplace=True)
df['EACP Appeals QA Sample'].fillna(0, inplace=True)
df['EACP Appeals Total Errors'].fillna(0, inplace=True)


convert_dict = {'EACP Appeals Tagged Volume':'int16', 'EACP Appeals QA Sample':'int16', 'EACP Appeals Total Errors':'int16',
                'EACP Appeals QA %':'float16','EACP Appeals Team Average':'float16'}

df = df.astype(convert_dict)




#converting accuracy columns into percentage by multiplying by 100
df['MCM QA %']= df['MCM QA %']*100
df['USN QA %']= df['USN QA %']*100
df['CCR QA %']= df['CCR QA %']*100
df['EIM Esc QA %']= df['EIM Esc QA %']*100
df['EIM FP QA %']= df['EIM FP QA %']*100
df['BO QA %']= df['BO QA %']*100

df['Transparency QA %']= df['Transparency QA %']*100
df['DCR QA %']= df['DCR QA %']*100
df['Coverage QA %']= df['Coverage QA %']*100
df['EPCR QA %']= df['EPCR QA %']*100
df['EACP QA %']= df['EACP QA %']*100
df['EACP Appeals QA %']= df['EACP Appeals QA %']*100

df['CCR Team Average']= df['CCR Team Average']*100
df['USN Team Average']= df['USN Team Average']*100
df['MCM Team Average']= df['MCM Team Average']*100
df['EIM Esc Team Average']= df['EIM Esc Team Average']*100
df['EIM FP Team Average']= df['EIM FP Team Average']*100
df['BO Team Average']= df['BO Team Average']*100
df['Transparency Team Average']= df['Transparency Team Average']*100
df['DCR Team Average']= df['DCR Team Average']*100
df['Coverage Team Average']= df['Coverage Team Average']*100
df['EPCR Team Average']= df['EPCR Team Average']*100
df['EACP Team Average']= df['EACP Team Average']*100
df['EACP Appeals Team Average']= df['EACP Appeals Team Average']*100




#adding weighted average column to the final report
   
df.loc[:,'Associate Goal(Weighted Average)']= ((df.loc[:, "USN Team Average"]*df.loc[:, "USN Tagged Volume"])+
                                               (df.loc[:, "MCM Team Average"]*df.loc[:, "MCM Tagged Volume"])+
                                               (df.loc[:, "CCR Team Average"]*df.loc[:, "CCR Tagged Volume"])+
                                               (df.loc[:, "EIM Esc Team Average"]*df.loc[:, "EIM Esc Tagged Volume"])+
                                               (df.loc[:, "EIM FP Team Average"]*df.loc[:, "EIM FP Tagged Volume"])+
                                               (df.loc[:, "BO Team Average"]*df.loc[:, "BO Tagged Volume"])+
                                               (df.loc[:, "Transparency Team Average"]*df.loc[:, "Transparency Tagged Volume"])+
                                               (df.loc[:, "DCR Team Average"]*df.loc[:, "DCR Tagged Volume"])+
                                               (df.loc[:, "Coverage Team Average"]*df.loc[:, "Coverage Tagged Volume"])+
                                               (df.loc[:, "EPCR Team Average"]*df.loc[:, "EPCR Tagged Volume"])+
                                               (df.loc[:, "EACP Team Average"]*df.loc[:, "EACP Tagged Volume"])+
                                               (df.loc[:, "EACP Appeals Team Average"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                                                  df.loc[:, "MCM Tagged Volume"]+
                                                                                                                                  df.loc[:, "CCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                                                  df.loc[:, "BO Tagged Volume"]+
                                                                                                                                  df.loc[:, "Transparency Tagged Volume"]+
                                                                                                                                  df.loc[:, "DCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "Coverage Tagged Volume"]+
                                                                                                                                  df.loc[:, "EPCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Appeals Tagged Volume"])





df['Associate Goal(Weighted Average)'].fillna(0, inplace=True)
#adding deviation in USN
df.loc[:,'USN Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "USN Tagged Volume"], df.loc[:, "USN QA %"],
                                                                         (df.loc[:, "USN QA %"]-df.loc[:, "USN Team Average"]))]

df.loc[:,'MCM Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "MCM Tagged Volume"], df.loc[:, "MCM QA %"],
                                                                         (df.loc[:, "MCM QA %"]-df.loc[:, "MCM Team Average"]))]

df.loc[:,'CCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "CCR Tagged Volume"], df.loc[:, "CCR QA %"],
                                                                         (df.loc[:, "CCR QA %"]-df.loc[:, "CCR Team Average"]))]

df.loc[:,'EIM Esc Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM Esc Tagged Volume"], df.loc[:, "EIM Esc QA %"],
                                                                         (df.loc[:, "EIM Esc QA %"]-df.loc[:, "EIM Esc Team Average"]))]

df.loc[:,'EIM FP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM FP Tagged Volume"], df.loc[:, "EIM FP QA %"],
                                                                         (df.loc[:, "EIM FP QA %"]-df.loc[:, "EIM FP Team Average"]))]

df.loc[:,'BO Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "BO Tagged Volume"], df.loc[:, "BO QA %"],
                                                                         (df.loc[:, "BO QA %"]-df.loc[:, "BO Team Average"]))]

df.loc[:,'Transparency Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Transparency Tagged Volume"], df.loc[:, "Transparency QA %"],
                                                                         (df.loc[:, "Transparency QA %"]-df.loc[:, "Transparency Team Average"]))]

df.loc[:,'DCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "DCR Tagged Volume"], df.loc[:, "DCR QA %"],
                                                                         (df.loc[:, "DCR QA %"]-df.loc[:, "DCR Team Average"]))]

df.loc[:,'Coverage Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Coverage Tagged Volume"], df.loc[:, "Coverage QA %"],
                                                                         (df.loc[:, "Coverage QA %"]-df.loc[:, "Coverage Team Average"]))]

df.loc[:,'EPCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EPCR Tagged Volume"], df.loc[:, "EPCR QA %"],
                                                                         (df.loc[:, "EPCR QA %"]-df.loc[:, "EPCR Team Average"]))]

df.loc[:,'EACP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Tagged Volume"], df.loc[:, "EACP QA %"],
                                                                         (df.loc[:, "EACP QA %"]-df.loc[:, "EACP Team Average"]))]

df.loc[:,'EACP Appeals Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Appeals Tagged Volume"], df.loc[:, "EACP Appeals QA %"],
                                                                         (df.loc[:, "EACP Appeals QA %"]-df.loc[:, "EACP Appeals Team Average"]))]






#net deviation column to the final report
df.loc[:,'Net Deviation']= ((df.loc[:, "USN Deviation"]*df.loc[:, "USN Tagged Volume"])+
                            (df.loc[:, "MCM Deviation"]*df.loc[:, "MCM Tagged Volume"])+
                            (df.loc[:, "CCR Deviation"]*df.loc[:, "CCR Tagged Volume"])+
                            (df.loc[:, "EIM Esc Deviation"]*df.loc[:, "EIM Esc Tagged Volume"])+
                            (df.loc[:, "EIM FP Deviation"]*df.loc[:, "EIM FP Tagged Volume"])+
                            (df.loc[:, "BO Deviation"]*df.loc[:, "BO Tagged Volume"])+
                            (df.loc[:, "Transparency Deviation"]*df.loc[:, "Transparency Tagged Volume"])+
                            (df.loc[:, "DCR Deviation"]*df.loc[:, "DCR Tagged Volume"])+
                            (df.loc[:, "Coverage Deviation"]*df.loc[:, "Coverage Tagged Volume"])+
                            (df.loc[:, "EPCR Deviation"]*df.loc[:, "EPCR Tagged Volume"])+
                            (df.loc[:, "EACP Deviation"]*df.loc[:, "EACP Tagged Volume"])+
                            (df.loc[:, "EACP Appeals Deviation"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                            df.loc[:, "MCM Tagged Volume"]+
                                                                                                            df.loc[:, "CCR Tagged Volume"]+
                                                                                                            df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                            df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                            df.loc[:, "BO Tagged Volume"]+
                                                                                                            df.loc[:, "Transparency Tagged Volume"]+
                                                                                                            df.loc[:, "DCR Tagged Volume"]+
                                                                                                            df.loc[:, "Coverage Tagged Volume"]+
                                                                                                            df.loc[:, "EPCR Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Appeals Tagged Volume"])



df['Net Deviation'].fillna(0, inplace=True)


#Final QA% (Weighed Average)
df.loc[:,'Final QA%(Weighted Average)']= [x if y==0 else z for x,y,z in zip(df.loc[:,'Associate Goal(Weighted Average)'],df.loc[:, "Net Deviation"],(df.loc[:,'Associate Goal(Weighted Average)']+ df.loc[:,'Net Deviation']))]

#converting NaN to 0

df['Final QA%(Weighted Average)'].fillna(0, inplace=True)



df= pd.merge(df, d1, how='left', on=['Login', 'Name'])




#getting columns position as per the report

df5= df[['Name', 'Login','Program','Manager','Shared Associate', 'Week','Year', 'Month',
         'USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
         'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
         'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
         'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
         'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
         'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
         'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
         'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
         'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
         'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
         'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
         'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
         'Associate Goal(Weighted Average)', 'USN Deviation',
         'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
         'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
         'EACP Deviation', 'EACP Appeals Deviation', 'Coverage Deviation','Net Deviation',
         'Final QA%(Weighted Average)']]

df5.loc[:, "Total QA Sample"]=[np.nan if x=="Yes" else y for x,y in zip(df5.loc[:, "Shared Associate"],(df5.loc[:, "USN QA Sample"]+df5.loc[:, "MCM QA Sample"]+df5.loc[:, "CCR QA Sample"]+
                                                                                                      df5.loc[:, "EIM Esc QA Sample"]+df5.loc[:, "EIM FP QA Sample"]+df5.loc[:, "BO QA Sample Audits"]+
                                                                                                      df5.loc[:, "Transparency QA Sample Audits"]+df5.loc[:, "DCR QA Sample"]+df5.loc[:, "EPCR QA Sample"]+
                                                                                                      df5.loc[:, "EACP QA Sample"]+df5.loc[:, "EACP Appeals QA Sample"]))]

df5.loc[:, "Total Errors"]= [np.nan if x=="Yes" else y for x,y in zip(df5.loc[:, "Shared Associate"],(df5.loc[:, "USN Total Errors"]+df5.loc[:, "MCM Total Errors"]+
                                                                                                    df5.loc[:, "CCR Total Errors"]+df5.loc[:, "EIM Esc Total Errors"]+
                                                                                                    df5.loc[:, "EIM FP Total Errors"]+
                                                                                                    df5.loc[:, "BO Total Errors"]+
                                                                                                    df5.loc[:, "Transparency Total Errors"]+df5.loc[:, "DCR Total Errors"]+
                                                                                                    df5.loc[:, "EPCR Total Errors"]+df5.loc[:, "EACP Appeals Total Errors"]))]


df5.loc[:, "QA % (Non Shared Associates)"]=[np.nan if x=='Yes' else y for x,y in zip (df5.loc[:, "Shared Associate"],
                                                                                   ((df5.loc[:, "Total QA Sample"]-df5.loc[:, "Total Errors"])/df5.loc[:,"Total QA Sample"]))]
        
df5.loc[:, "QA % (Non Shared Associates)"]= df5.loc[:, 'QA % (Non Shared Associates)']*100

df5.loc[:, "Final QA%(Weighted Average)"]=[np.nan if x=='No' else y for x,y in zip (df5.loc[:, "Shared Associate"],df5.loc[:, "Final QA%(Weighted Average)"])]
        
cols=['USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
      'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
      'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
      'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
      'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
      'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
      'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
      'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
      'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
      'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
      'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
      'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
      'Associate Goal(Weighted Average)', 'USN Deviation',
      'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
      'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
      'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation','Total QA Sample', 'Total Errors',
      ]

df5[cols].replace(to_replace= 0, value= np.nan, inplace= True)

        

df5.sort_values(by=["Week", 'Login'], inplace=True)


#adding productivity to the report
df5= pd.merge(df5, tracker, how='left', on=['Login', 'Week'])

df5= df5[['Name', 'Login','Program','Manager','Shared Associate', 'Week','Year', 'Month',
          'USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
          'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
          'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
          'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
          'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
          'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
          'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
          'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
          'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
          'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
          'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
          'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
          'Associate Goal(Weighted Average)', 'USN Deviation',
          'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
          'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
          'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation',
          'Final QA%(Weighted Average)','Total QA Sample', 'Total Errors', 'QA % (Non Shared Associates)',
          'Total','Actual Cumputed Hours','Actual Cumputed NPT',
          'Leave Hours','Weighted Productive Sum','Weighted Productivity %']]



df5['Total'].replace(to_replace= 0, value= np.nan, inplace= True)
df5['Actual Cumputed Hours'].replace(to_replace= 0, value= np.nan, inplace= True)
df5['Actual Cumputed NPT'].replace(to_replace= 0, value= np.nan, inplace= True)
df5['Leave Hours'].replace(to_replace= 0, value= np.nan, inplace= True)
df5['Weighted Productive Sum'].replace(to_replace= 0, value= np.nan, inplace= True)
df5['Weighted Productivity %'].replace(to_replace= 0, value= np.nan, inplace= True)

df5= df5.round({"Final QA%(Weighted Average)":2, "Weighted Productivity %":2, "QA % (Non Shared Associates) %":2})
        

# def append_df_to_excel(filename, final_report, sheet_name, startrow=None,
#                            truncate_sheet= False, 
#                            **to_excel_kwargs):
    
#         # Excel file doesn't exist - saving and exiting
#         if not os.path.isfile(filename):
#             final_report.to_excel(
#                 filename,
#                 sheet_name=sheet_name, 
#                 startrow=startrow if startrow is not None else 0, 
#                 header=True,
#                 index=False,
#                 **to_excel_kwargs)
#             return
#         # ignore [engine] parameter if it was passed
#         if 'engine' in to_excel_kwargs:
#             to_excel_kwargs.pop('engine')
    
#         writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
    
#         # try to open an existing workbook
#         writer.book = load_workbook(filename)
        
#         # get the last row in the existing Excel sheet
#         # if it was not specified explicitly
#         if startrow is None and sheet_name in writer.book.sheetnames:
#             startrow = writer.book[sheet_name].max_row
    
#         # truncate sheet
#         if truncate_sheet and sheet_name in writer.book.sheetnames:
#             # index of [sheet_name] sheet
#             idx = writer.book.sheetnames.index(sheet_name)
#             # remove [sheet_name]
#             writer.book.remove(writer.book.worksheets[idx])
#             # create an empty sheet [sheet_name] using old index
#             writer.book.create_sheet(sheet_name, idx)
        
#         # copy existing sheets
#         writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    
#         if startrow is None:
#             startrow = 0
    
#         # write out the new sheet
#         final_report.to_excel(writer, sheet_name, startrow=startrow,header=False,index=False, **to_excel_kwargs)
    
#         # save the workbook
#         writer.save()
        
# append_df_to_excel(path1+'/Final_Report/Master_Report_'+str(year_data)+'.xlsx',df5, 'Master_data')

# path=r'C:/Users/nkcho/Desktop/Final_Report'

# filename1=r'C:/Users/nkcho/Desktop/Final_Report/Master_Report_2021.xlsx'




# data=pd.read_excel(filename1, sheet_name=0)
# data=pd.DataFrame(data)


data=df5.copy()

files = (path1 +'/Master_Login.xlsx')
        
d1 = pd.read_excel(files, sheet_name=0)

d1=pd.DataFrame(d1)
        

consolidated_weekly= data[['Name','Login','Program','Manager','Shared Associate','Week','Year','Month',
                    'Final QA%(Weighted Average)','Total QA Sample', 
                    'Total Errors', 'QA % (Non Shared Associates)', 'Weighted Productivity %']]


#MONTHLY REPORT
#define the sorter which is Month used in order to sort the data as per month

Month_sorter= ['January', 'February', 'March', 'April', 'May', 'June', 
               'July', 'August', 'September', 'October', 'November', 'December']


data_1= data.copy()
data_1.reset_index(drop=True, inplace=True)
#converting data into dataframe
data_1= pd.DataFrame(data_1)

#converting NA to 0 in month column

data_1.replace(to_replace= np.nan, value= 0, inplace= True)


#creating a new dataframe in the same sheet
data4 = data_1.groupby(['Month', 'Login'])[['USN Tagged Volume','MCM Tagged Volume', 'CCR Tagged Volume', 'EIM Esc Tagged Volume',
                                            'EIM FP Tagged Volume','BO Tagged Volume',
                                            'EPCR Tagged Volume','EACP Tagged Volume','EACP Appeals Tagged Volume',
                                            'DCR Tagged Volume', 'Coverage Tagged Volume',
                                            'Transparency Tagged Volume','USN Total Errors','USN QA Sample',
                                            'MCM Total Errors','MCM QA Sample','CCR Total Errors','CCR QA Sample',
                                            'EIM Esc Total Errors', 'EIM Esc QA Sample','EIM FP Total Errors', 'EIM FP QA Sample',
                                            'BO QA Sample Audits','BO Total Errors',
                                            'Transparency QA Sample Audits', 'Transparency Total Errors',
                                            'DCR QA Sample', 'DCR Total Errors','Coverage QA Sample', 'Coverage Total Errors',
                                            'EPCR QA Sample', 'EPCR Total Errors','EACP QA Sample', 'EACP Total Errors',
                                            'EACP Appeals QA Sample', 'EACP Appeals Total Errors','Total QA Sample', 'Total Errors',
                                            'Total','Actual Cumputed Hours','Actual Cumputed NPT',
                                            'Leave Hours','Weighted Productive Sum']].sum()


data4= pd.DataFrame(data4)
data4= data4.reset_index()

data4.loc[:,'USN QA %']= (data4.loc[:, "USN QA Sample"]
                          -data4.loc[:,"USN Total Errors"])/data4.loc[:,"USN QA Sample"]
data4.loc[:,'MCM QA %']= (data4.loc[:, "MCM QA Sample"]
                          -data4.loc[:,"MCM Total Errors"])/data4.loc[:,"MCM QA Sample"]
data4.loc[:,'CCR QA %']= (data4.loc[:, "CCR QA Sample"]
                          -data4.loc[:,"CCR Total Errors"])/data4.loc[:,"CCR QA Sample"]
data4.loc[:,'EIM Esc QA %']= (data4.loc[:, "EIM Esc QA Sample"]
                          -data4.loc[:,"EIM Esc Total Errors"])/data4.loc[:,"EIM Esc QA Sample"]
data4.loc[:,'EIM FP QA %']= (data4.loc[:, "EIM FP QA Sample"]
                          -data4.loc[:,"EIM FP Total Errors"])/data4.loc[:,"EIM FP QA Sample"]
data4.loc[:,'BO QA %']= (data4.loc[:, "BO QA Sample Audits"]
                          -data4.loc[:,"BO Total Errors"])/data4.loc[:,"BO QA Sample Audits"]
data4.loc[:,'Transparency QA %']= (data4.loc[:, "Transparency QA Sample Audits"]
                          -data4.loc[:,"Transparency Total Errors"])/data4.loc[:,"Transparency QA Sample Audits"]
data4.loc[:,'EPCR QA %']= (data4.loc[:, "EPCR QA Sample"]
                          -data4.loc[:,"EPCR Total Errors"])/data4.loc[:,"EPCR QA Sample"]
data4.loc[:,'EACP QA %']= (data4.loc[:, "EACP QA Sample"]
                          -data4.loc[:,"EACP Total Errors"])/data4.loc[:,"EACP QA Sample"]
data4.loc[:,'EACP Appeals QA %']= (data4.loc[:, "EACP Appeals QA Sample"]
                                   -data4.loc[:,"EACP Appeals Total Errors"])/data4.loc[:,"EACP Appeals QA Sample"]
data4.loc[:,'DCR QA %']= (data4.loc[:, "DCR QA Sample"]
                          -data4.loc[:,"DCR Total Errors"])/data4.loc[:,"DCR QA Sample"]
data4.loc[:,'Coverage QA %']= (data4.loc[:, "Coverage QA Sample"]
                          -data4.loc[:,"Coverage Total Errors"])/data4.loc[:,"Coverage QA Sample"]


#converting accuracy columns into percentage by multiplying by 100
data4['MCM QA %']= data4['MCM QA %']*100
data4['USN QA %']= data4['USN QA %']*100
data4['CCR QA %']= data4['CCR QA %']*100
data4['EIM Esc QA %']= data4['EIM Esc QA %']*100
data4['EIM FP QA %']= data4['EIM FP QA %']*100
data4['BO QA %']= data4['BO QA %']*100

data4['Transparency QA %']= data4['Transparency QA %']*100
data4['EPCR QA %']= data4['EPCR QA %']*100
data4['EACP QA %']= data4['EACP QA %']*100
data4['EACP Appeals QA %']= data4['EACP Appeals QA %']*100
data4['DCR QA %']= data4['DCR QA %']*100
data4['Coverage QA %']= data4['Coverage QA %']*100

#sort values based on month_sorter in the final data
data4.Month= data4.Month.astype('category')

data4.Month.cat.set_categories(Month_sorter, inplace=True)

data4.sort_values(by=['Month', 'Login'], inplace=True)

#converting category to str in month column
data4 = data4.astype({"Month": 'str'})

#monthly team average

#USN

ta_usno= ta_usno.rename(columns={ta_usno.columns[0]:'Week'})
#ta_usno= ta_usno.drop(ta_usno[(ta_usno.loc[:, "Week"]!=w1)].index)

ta_usn1= ta_usno[ta_usno.columns[(0,7,9), ]].copy()

ta_usn1=pd.merge(ta_usn1, d2, on=['Week'])
ta_usn1= ta_usn1.groupby(['Month'])[['TMM USN Sample','Correct Votes']].sum()

ta_usn1.reset_index(inplace=True)
ta_usn1.loc[:, 'USN Team Average']= (ta_usn1.iloc[:, 2]/ta_usn1.iloc[:, 1])*100

ta_usn1= ta_usn1[['Month', 'USN Team Average']]

data4= pd.merge(data4, ta_usn1, how= 'left', on=['Month'])

data4.loc[:, "USN Team Average"].fillna(0, inplace=True)
#MCM


#getting team average for MCM
team_average= mcmo.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= mcmo.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)
team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])
team_average= team_average.groupby(['Month'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Month'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg_mcm= pd.merge(team_average, team_average1, on=['Month'])
team_avg_mcm= pd.DataFrame(team_avg_mcm)
team_avg=team_avg_mcm.copy()

team_avg.loc[:,'MCM Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'MCM Team Average']] 

data4= pd.merge(data4, team_avg, how= 'left', on=['Month'])
data4.loc[:, "MCM Team Average"].fillna(0, inplace=True)

#CCR

#getting relevant columns for TA
ta_ccr1= ta_ccro[ta_ccro.columns[(0,8,9), ]].copy()
ta_ccr1= ta_ccr1.rename(columns={ta_ccr1.columns[0]:'Week'})
#ta_ccr1= ta_ccr1.drop(ta_ccr1[(ta_ccr1.loc[:, "Week"]!=w1)].index)

ta_ccr1=pd.merge(ta_ccr1, d2, on=['Week'])
ta_ccr1= ta_ccr1.groupby(['Month'])[['Correct Votes','Traditional CCR Sample']].sum()

ta_ccr1.reset_index(inplace=True)
ta_ccr1.loc[:, 'CCR Team Average']= (ta_ccr1.iloc[:, 1]/ ta_ccr1.iloc[:, 2])*100
ta_ccr1= ta_ccr1[['Month','CCR Team Average']]


data4= pd.merge(data4, ta_ccr1, how= 'left', on=['Month'])
data4.loc[:, "CCR Team Average"].fillna(0, inplace=True)

#EIM Escalations

#getting team average for EIM Esc
team_avg= eimo.groupby(['Week'])[['EIM Esc QA Sample','EIM Esc Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)

team_avg=pd.merge(team_avg, d2, how='left', on=['Week'])

team_avg= team_avg.groupby(['Month'])[['EIM Esc QA Sample','EIM Esc Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EIM Esc Team Average']= ((team_avg['EIM Esc QA Sample']- team_avg['EIM Esc Total Errors'])/team_avg['EIM Esc QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'EIM Esc Team Average']]
data4= pd.merge(data4, team_avg, how= 'left', on=['Month'])

data4.loc[:, "EIM Esc Team Average"].fillna(0, inplace=True)
#EIM FP

#getting team average for EIM FP
team_avg= eim_o.groupby(['Week'])[['EIM FP QA Sample','EIM FP Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)
team_avg=pd.merge(team_avg, d2, how='left', on=['Week'])

team_avg= team_avg.groupby(['Month'])[['EIM FP QA Sample','EIM FP Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EIM FP Team Average']= ((team_avg['EIM FP QA Sample']- team_avg['EIM FP Total Errors'])/team_avg['EIM FP QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'EIM FP Team Average']]

data4= pd.merge(data4, team_avg, how= 'left', on=['Month'])
data4.loc[:, "EIM FP Team Average"].fillna(0, inplace=True)

#Adhoc
adhoc2=adhoco[['Week', 'BO Team Average', 'Transparency Team Average']]

adhoc2=pd.merge(adhoc2, d2, on=['Week'], how='left')

adhoc2=adhoc2.groupby(['Month'])[['BO Team Average', 'Transparency Team Average']].mean()
adhoc2.reset_index(inplace=True)

adhoc2=adhoc2[['Month','BO Team Average', 'Transparency Team Average']]

adhoc2.loc[:, "BO Team Average"]=adhoc2.loc[:, "BO Team Average"]*100
adhoc2.loc[:, "Transparency Team Average"]=adhoc2.loc[:, "Transparency Team Average"]*100

data4=pd.merge(data4, adhoc2, how='left',on=['Month'])

data4.loc[:, "BO Team Average"].fillna(0, inplace=True)
data4.loc[:, "Transparency Team Average"].fillna(0, inplace=True)

#DCR

#getting team average for MCM

team_average= dcro.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= dcro.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Month'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Month'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Month'])
team_avg= pd.DataFrame(team_avg)



team_avg.loc[:,'DCR Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'DCR Team Average']]


data4=pd.merge(data4, team_avg, how='left',on=['Month'])
data4.loc[:, "DCR Team Average"].fillna(0, inplace=True)


#Coverage

#getting team average for Coverage
team_average= coverageo.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= coverageo.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Month'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Month'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)
team_avg= pd.merge(team_average, team_average1, on=['Month'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'Coverage Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'Coverage Team Average']]

data4=pd.merge(data4, team_avg, how='left',on=['Month'])
data4.loc[:, "Coverage Team Average"].fillna(0, inplace=True)


#EPCR
#getting team average for EPCR
team_average= epcr1.groupby(['Week'])[['EPCR QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= epcr1.groupby(['Week'])[['EPCR Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)
team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Month'])[['EPCR QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Month'])[['EPCR Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Month'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EPCR Team Average']= ((team_avg['EPCR QA Sample']- team_avg['EPCR Total Errors'])/team_avg['EPCR QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'EPCR Team Average']]

data4=pd.merge(data4, team_avg, how='left',on=['Month'])
data4.loc[:, "EPCR Team Average"].fillna(0, inplace=True)

#EACP appeals

#getting team average for EPCR
team_average= eacp_appeals.groupby(['Week'])[['EACP Appeals QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp_appeals.groupby(['Week'])[['EACP Appeals Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Month'])[['EACP Appeals QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Month'])[['EACP Appeals Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Month'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Appeals Team Average']= ((team_avg['EACP Appeals QA Sample']- team_avg['EACP Appeals Total Errors'])/team_avg['EACP Appeals QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'EACP Appeals Team Average']]

data4=pd.merge(data4, team_avg, how='left',on=['Month'])
data4.loc[:, "EACP Appeals Team Average"].fillna(0, inplace=True)

#EACP
#getting team average for EPCR
team_average= eacp1.groupby(['Week'])[['EACP QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp1.groupby(['Week'])[['EACP Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Month'])[['EACP QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Month'])[['EACP Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Month'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Team Average']= ((team_avg['EACP QA Sample']- team_avg['EACP Total Errors'])/team_avg['EACP QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Month', 'EACP Team Average']]

data4=pd.merge(data4, team_avg, how='left',on=['Month'])
data4.loc[:, "EACP Team Average"].fillna(0, inplace=True)





df=data4.copy()

#adding weighted average column to the final report

#adding weighted average column to the final report
   
df.loc[:,'Associate Goal(Weighted Average)']= ((df.loc[:, "USN Team Average"]*df.loc[:, "USN Tagged Volume"])+
                                               (df.loc[:, "MCM Team Average"]*df.loc[:, "MCM Tagged Volume"])+
                                               (df.loc[:, "CCR Team Average"]*df.loc[:, "CCR Tagged Volume"])+
                                               (df.loc[:, "EIM Esc Team Average"]*df.loc[:, "EIM Esc Tagged Volume"])+
                                               (df.loc[:, "EIM FP Team Average"]*df.loc[:, "EIM FP Tagged Volume"])+
                                               (df.loc[:, "BO Team Average"]*df.loc[:, "BO Tagged Volume"])+
                                               (df.loc[:, "Transparency Team Average"]*df.loc[:, "Transparency Tagged Volume"])+
                                               (df.loc[:, "DCR Team Average"]*df.loc[:, "DCR Tagged Volume"])+
                                               (df.loc[:, "Coverage Team Average"]*df.loc[:, "Coverage Tagged Volume"])+
                                               (df.loc[:, "EPCR Team Average"]*df.loc[:, "EPCR Tagged Volume"])+
                                               (df.loc[:, "EACP Team Average"]*df.loc[:, "EACP Tagged Volume"])+
                                               (df.loc[:, "EACP Appeals Team Average"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                                                  df.loc[:, "MCM Tagged Volume"]+
                                                                                                                                  df.loc[:, "CCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                                                  df.loc[:, "BO Tagged Volume"]+
                                                                                                                                  df.loc[:, "Transparency Tagged Volume"]+
                                                                                                                                  df.loc[:, "DCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "Coverage Tagged Volume"]+
                                                                                                                                  df.loc[:, "EPCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Appeals Tagged Volume"])





df['Associate Goal(Weighted Average)'].fillna(0, inplace=True)
#adding deviation in USN
df.loc[:,'USN Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "USN Tagged Volume"], df.loc[:, "USN QA %"],
                                                                         (df.loc[:, "USN QA %"]-df.loc[:, "USN Team Average"]))]

df.loc[:,'MCM Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "MCM Tagged Volume"], df.loc[:, "MCM QA %"],
                                                                         (df.loc[:, "MCM QA %"]-df.loc[:, "MCM Team Average"]))]

df.loc[:,'CCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "CCR Tagged Volume"], df.loc[:, "CCR QA %"],
                                                                         (df.loc[:, "CCR QA %"]-df.loc[:, "CCR Team Average"]))]

df.loc[:,'EIM Esc Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM Esc Tagged Volume"], df.loc[:, "EIM Esc QA %"],
                                                                         (df.loc[:, "EIM Esc QA %"]-df.loc[:, "EIM Esc Team Average"]))]

df.loc[:,'EIM FP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM FP Tagged Volume"], df.loc[:, "EIM FP QA %"],
                                                                         (df.loc[:, "EIM FP QA %"]-df.loc[:, "EIM FP Team Average"]))]

df.loc[:,'BO Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "BO Tagged Volume"], df.loc[:, "BO QA %"],
                                                                         (df.loc[:, "BO QA %"]-df.loc[:, "BO Team Average"]))]

df.loc[:,'Transparency Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Transparency Tagged Volume"], df.loc[:, "Transparency QA %"],
                                                                         (df.loc[:, "Transparency QA %"]-df.loc[:, "Transparency Team Average"]))]

df.loc[:,'DCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "DCR Tagged Volume"], df.loc[:, "DCR QA %"],
                                                                         (df.loc[:, "DCR QA %"]-df.loc[:, "DCR Team Average"]))]

df.loc[:,'Coverage Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Coverage Tagged Volume"], df.loc[:, "Coverage QA %"],
                                                                         (df.loc[:, "Coverage QA %"]-df.loc[:, "Coverage Team Average"]))]

df.loc[:,'EPCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EPCR Tagged Volume"], df.loc[:, "EPCR QA %"],
                                                                         (df.loc[:, "EPCR QA %"]-df.loc[:, "EPCR Team Average"]))]

df.loc[:,'EACP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Tagged Volume"], df.loc[:, "EACP QA %"],
                                                                         (df.loc[:, "EACP QA %"]-df.loc[:, "EACP Team Average"]))]

df.loc[:,'EACP Appeals Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Appeals Tagged Volume"], df.loc[:, "EACP Appeals QA %"],
                                                                         (df.loc[:, "EACP Appeals QA %"]-df.loc[:, "EACP Appeals Team Average"]))]






#net deviation column to the final report
df.loc[:,'Net Deviation']= ((df.loc[:, "USN Deviation"]*df.loc[:, "USN Tagged Volume"])+
                            (df.loc[:, "MCM Deviation"]*df.loc[:, "MCM Tagged Volume"])+
                            (df.loc[:, "CCR Deviation"]*df.loc[:, "CCR Tagged Volume"])+
                            (df.loc[:, "EIM Esc Deviation"]*df.loc[:, "EIM Esc Tagged Volume"])+
                            (df.loc[:, "EIM FP Deviation"]*df.loc[:, "EIM FP Tagged Volume"])+
                            (df.loc[:, "BO Deviation"]*df.loc[:, "BO Tagged Volume"])+
                            (df.loc[:, "Transparency Deviation"]*df.loc[:, "Transparency Tagged Volume"])+
                            (df.loc[:, "DCR Deviation"]*df.loc[:, "DCR Tagged Volume"])+
                            (df.loc[:, "Coverage Deviation"]*df.loc[:, "Coverage Tagged Volume"])+
                            (df.loc[:, "EPCR Deviation"]*df.loc[:, "EPCR Tagged Volume"])+
                            (df.loc[:, "EACP Deviation"]*df.loc[:, "EACP Tagged Volume"])+
                            (df.loc[:, "EACP Appeals Deviation"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                            df.loc[:, "MCM Tagged Volume"]+
                                                                                                            df.loc[:, "CCR Tagged Volume"]+
                                                                                                            df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                            df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                            df.loc[:, "BO Tagged Volume"]+
                                                                                                            df.loc[:, "Transparency Tagged Volume"]+
                                                                                                            df.loc[:, "DCR Tagged Volume"]+
                                                                                                            df.loc[:, "Coverage Tagged Volume"]+
                                                                                                            df.loc[:, "EPCR Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Appeals Tagged Volume"])



df['Net Deviation'].fillna(0, inplace=True)
#Final QA% (Weighed Average)
df.loc[:,'Final QA%(Weighted Average)']= [x if y==0 else z for x,y,z in zip(df.loc[:,'Associate Goal(Weighted Average)'],df.loc[:, "Net Deviation"],(df.loc[:,'Associate Goal(Weighted Average)']+ df.loc[:,'Net Deviation']))]

#converting NaN to 0

df['Final QA%(Weighted Average)'].fillna(0, inplace=True)



df= pd.merge(df, d1, how='left', on=['Login'])


df.loc[:, "Year"]=year_data

#getting columns position as per the report

df5= df[['Name','Login','Program','Manager','Shared Associate','Year', 'Month',
         'USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
         'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
         'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
         'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
         'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
         'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
         'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
         'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
         'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
         'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
         'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
         'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
         'Associate Goal(Weighted Average)', 'USN Deviation',
         'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
         'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
         'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation',
         'Final QA%(Weighted Average)',
         'Total','Actual Cumputed Hours','Actual Cumputed NPT',
         'Leave Hours','Weighted Productive Sum']]

df5.loc[:, "Total QA Sample"]=[np.nan if x=="Yes" else y for x,y in zip(df5.loc[:, "Shared Associate"],(df5.loc[:, "USN QA Sample"]+df5.loc[:, "MCM QA Sample"]+df5.loc[:, "CCR QA Sample"]+
                                                                                                      df5.loc[:, "EIM Esc QA Sample"]+df5.loc[:, "EIM FP QA Sample"]+df5.loc[:, "BO QA Sample Audits"]+
                                                                                                      df5.loc[:, "Transparency QA Sample Audits"]+df5.loc[:, "DCR QA Sample"]+df5.loc[:, "EPCR QA Sample"]+
                                                                                                      df5.loc[:, "EACP QA Sample"]+df5.loc[:, "EACP Appeals QA Sample"]+df5.loc[:, "Coverage QA Sample"]))]

df5.loc[:, "Total Errors"]= [np.nan if x=="Yes" else y for x,y in zip(df5.loc[:, "Shared Associate"],(df5.loc[:, "USN Total Errors"]+df5.loc[:, "MCM Total Errors"]+
                                                                                                    df5.loc[:, "CCR Total Errors"]+df5.loc[:, "EIM Esc Total Errors"]+
                                                                                                    df5.loc[:, "EIM FP Total Errors"]+df5.loc[:, "Coverage Total Errors"]+
                                                                                                    df5.loc[:, "BO Total Errors"]+
                                                                                                    df5.loc[:, "Transparency Total Errors"]+df5.loc[:, "DCR Total Errors"]+
                                                                                                    df5.loc[:, "EPCR Total Errors"]+df5.loc[:, "EACP Appeals Total Errors"]+df5.loc[:, "EACP Total Errors"]))]


df5.loc[:, "QA % (Non Shared Associates)"]=[np.nan if x=='Yes' else y for x,y in zip (df5.loc[:, "Shared Associate"],
                                                                                   ((df5.loc[:, "Total QA Sample"]-df5.loc[:, "Total Errors"])/df5.loc[:,"Total QA Sample"]))]
        
df5.loc[:, "QA % (Non Shared Associates)"]= df5.loc[:, 'QA % (Non Shared Associates)']*100

df5.loc[:, "Final QA%(Weighted Average)"]=[np.nan if x=='No' else y for x,y in zip (df5.loc[:, "Shared Associate"],df5.loc[:, "Final QA%(Weighted Average)"])]
        
df5.loc[:, "Weighted Productivity %"]= [0 if x==0 else y for x,y in zip(df5.loc[:, "Actual Cumputed Hours"],(df5.loc[:, "Weighted Productive Sum"]/df5.loc[:, "Actual Cumputed Hours"]))]
df5.loc[:, "Weighted Productivity %"]=df5.loc[:, "Weighted Productivity %"]*100


cols=['USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
      'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
      'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
      'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
      'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
      'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
      'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
      'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
      'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
      'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
      'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
      'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
      'Associate Goal(Weighted Average)', 'USN Deviation',
      'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
      'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
      'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation',
      'Total','Actual Cumputed Hours','Actual Cumputed NPT',
      'Leave Hours','Weighted Productive Sum', 'Weighted Productivity %']

df5[cols].replace(to_replace= 0, value= np.nan, inplace= True)

cols=['Total','Actual Cumputed Hours','Actual Cumputed NPT',
      'Leave Hours','Weighted Productive Sum','Weighted Productivity %']

df5[cols].replace(to_replace= 0, value= np.nan, inplace= True)

df5= df5.round({"Final QA%(Weighted Average)":2, "Weighted Productivity %":2, "QA % (Non Shared Associates) %":2})
        
        
df5=df5[['Name','Login','Program','Manager','Shared Associate','Year', 'Month',
         'USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
         'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
         'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
         'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
         'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
         'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
         'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
         'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
         'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
         'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
         'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
         'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
         'Associate Goal(Weighted Average)', 'USN Deviation',
         'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
         'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
         'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation',
         'Final QA%(Weighted Average)','Total QA Sample', 'Total Errors','QA % (Non Shared Associates)',
         'Total','Actual Cumputed Hours','Actual Cumputed NPT',
         'Leave Hours','Weighted Productive Sum', 'Weighted Productivity %']]


    
monthly_consolidated= df5[['Name','Login','Program','Manager','Shared Associate','Year','Month',
                             'Final QA%(Weighted Average)','Total QA Sample', 
                             'Total Errors', 'QA % (Non Shared Associates)', 'Weighted Productivity %']].copy()
#QUARTERLY REPORT

data_5=df5.copy()

qtr1= data_5.copy()
qtr1.reset_index(drop=True, inplace=True)
qtr1.replace([np.inf, -np.inf], np.nan, inplace=True)
qtr1= qtr1.fillna(0)

#converting category to str in month column
qtr1 = qtr1.astype({"Month": str,
                    'Transparency QA Sample Audits':int})

#removing NA from month
qtr1.drop(qtr1[qtr1['Month']=="NA"].index, inplace= True)

qtr_data=d2[['Month', 'Qtr']]

qtr1=pd.merge(qtr1,qtr_data,how='left', on=['Month'])

qtr_1= qtr1.copy()

qtr_final1 = qtr_1.groupby(['Qtr', 'Login'])[['USN Tagged Volume','MCM Tagged Volume', 'CCR Tagged Volume', 'EIM Esc Tagged Volume',
                                              'EIM FP Tagged Volume','BO Tagged Volume',
                                              'EPCR Tagged Volume','EACP Tagged Volume','EACP Appeals Tagged Volume',
                                              'DCR Tagged Volume', 'Coverage Tagged Volume',
                                              'Transparency Tagged Volume','USN Total Errors','USN QA Sample',
                                              'MCM Total Errors','MCM QA Sample','CCR Total Errors','CCR QA Sample',
                                              'EIM Esc Total Errors', 'EIM Esc QA Sample','EIM FP Total Errors', 'EIM FP QA Sample',
                                              'BO QA Sample Audits','BO Total Errors',
                                              'Transparency QA Sample Audits', 'Transparency Total Errors',
                                              'DCR QA Sample', 'DCR Total Errors','Coverage QA Sample', 'Coverage Total Errors',
                                              'EPCR QA Sample', 'EPCR Total Errors','EACP QA Sample', 'EACP Total Errors',
                                              'EACP Appeals QA Sample', 'EACP Appeals Total Errors','Total QA Sample', 'Total Errors',
                                              'Total','Actual Cumputed Hours','Actual Cumputed NPT',
                                              'Leave Hours','Weighted Productive Sum']].sum()

qtr_final1 = qtr_final1.astype({'Transparency QA Sample Audits':float})
qtr_final1= qtr_final1.reset_index()

qtr_final1= pd.DataFrame(qtr_final1)

qtr_final1.replace([np.inf, -np.inf], np.nan, inplace=True)
qtr_final1= qtr_final1.fillna(0)


qtr_final1.loc[:,'USN QA %']= (qtr_final1.loc[:, "USN QA Sample"]
                               -qtr_final1.loc[:,"USN Total Errors"])/qtr_final1.loc[:,"USN QA Sample"]
qtr_final1.loc[:,'MCM QA %']= (qtr_final1.loc[:, "MCM QA Sample"]
                               -qtr_final1.loc[:,"MCM Total Errors"])/qtr_final1.loc[:,"MCM QA Sample"]
qtr_final1.loc[:,'CCR QA %']= (qtr_final1.loc[:, "CCR QA Sample"]
                               -qtr_final1.loc[:,"CCR Total Errors"])/qtr_final1.loc[:,"CCR QA Sample"]
qtr_final1.loc[:,'EIM Esc QA %']= (qtr_final1.loc[:, "EIM Esc QA Sample"]
                                   -qtr_final1.loc[:,"EIM Esc Total Errors"])/qtr_final1.loc[:,"EIM Esc QA Sample"]
qtr_final1.loc[:,'EIM FP QA %']= (qtr_final1.loc[:, "EIM FP QA Sample"]
                                  -qtr_final1.loc[:,"EIM FP Total Errors"])/qtr_final1.loc[:,"EIM FP QA Sample"]
qtr_final1.loc[:,'BO QA %']= (qtr_final1.loc[:, "BO QA Sample Audits"]
                              -qtr_final1.loc[:,"BO Total Errors"])/qtr_final1.loc[:,"BO QA Sample Audits"]
qtr_final1.loc[:,'Transparency QA %']= (qtr_final1.loc[:, "Transparency QA Sample Audits"]
                                        -qtr_final1.loc[:,"Transparency Total Errors"])/qtr_final1.loc[:,"Transparency QA Sample Audits"]
qtr_final1.loc[:,'EPCR QA %']= (qtr_final1.loc[:, "EPCR QA Sample"]
                                -qtr_final1.loc[:,"EPCR Total Errors"])/qtr_final1.loc[:,"EPCR QA Sample"]
qtr_final1.loc[:,'EACP QA %']= (qtr_final1.loc[:, "EACP QA Sample"]
                                -qtr_final1.loc[:,"EACP Total Errors"])/qtr_final1.loc[:,"EACP QA Sample"]
qtr_final1.loc[:,'EACP Appeals QA %']= (qtr_final1.loc[:, "EACP Appeals QA Sample"]
                                        -qtr_final1.loc[:,"EACP Appeals Total Errors"])/qtr_final1.loc[:,"EACP Appeals QA Sample"]
qtr_final1.loc[:,'DCR QA %']= (qtr_final1.loc[:, "DCR QA Sample"]
                               -qtr_final1.loc[:,"DCR Total Errors"])/qtr_final1.loc[:,"DCR QA Sample"]
qtr_final1.loc[:,'Coverage QA %']= (qtr_final1.loc[:, "Coverage QA Sample"]
                                    -qtr_final1.loc[:,"Coverage Total Errors"])/qtr_final1.loc[:,"Coverage QA Sample"]


qtr_final1.replace([np.inf, -np.inf], np.nan, inplace=True)
qtr_final1= qtr_final1.fillna(0)

#converting accuracy columns into percentage by multiplying by 100
qtr_final1['MCM QA %']= qtr_final1['MCM QA %']*100
qtr_final1['USN QA %']= qtr_final1['USN QA %']*100
qtr_final1['CCR QA %']= qtr_final1['CCR QA %']*100
qtr_final1['EIM Esc QA %']= qtr_final1['EIM Esc QA %']*100
qtr_final1['EIM FP QA %']= qtr_final1['EIM FP QA %']*100
qtr_final1['BO QA %']= qtr_final1['BO QA %']*100
qtr_final1['Transparency QA %']= qtr_final1['Transparency QA %']*100
qtr_final1['EPCR QA %']= qtr_final1['EPCR QA %']*100
qtr_final1['EACP QA %']= qtr_final1['EACP QA %']*100
qtr_final1['EACP Appeals QA %']= qtr_final1['EACP Appeals QA %']*100
qtr_final1['DCR QA %']= qtr_final1['DCR QA %']*100
qtr_final1['Coverage QA %']= qtr_final1['Coverage QA %']*100


#quarterly team average

#USN

ta_usno= ta_usno.rename(columns={ta_usno.columns[0]:'Week'})
#ta_usno= ta_usno.drop(ta_usno[(ta_usno.loc[:, "Week"]!=w1)].index)

ta_usn1= ta_usno[ta_usno.columns[(0,7,9), ]].copy()




ta_usn1=pd.merge(ta_usn1, d2, on=['Week'])
ta_usn1= ta_usn1.groupby(['Qtr'])[['TMM USN Sample','Correct Votes']].sum()
ta_usn1.reset_index(inplace=True)

ta_usn1.loc[:, 'USN Team Average']= (ta_usn1.iloc[:, 2]/ta_usn1.iloc[:, 1])*100

ta_usn1= ta_usn1[['Qtr', 'USN Team Average']]




qtr_final3= ta_usn1.copy()

qtr_final3.loc[:, "USN Team Average"].fillna(0, inplace=True)
#MCM


#getting team average for MCM
team_average= mcmo.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= mcmo.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)
team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])
team_average= team_average.groupby(['Qtr'])[['QA Sample']].sum()
team_average= team_average.reset_index()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Qtr'])[['Total Errors']].sum()
team_average1= team_average1.reset_index()
team_average1= pd.DataFrame(team_average1)

team_avg_mcm= pd.merge(team_average, team_average1, on=['Qtr'])
team_avg_mcm= pd.DataFrame(team_avg_mcm)
team_avg=team_avg_mcm.copy()

team_avg.loc[:,'MCM Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'MCM Team Average']] 

qtr_final3= pd.merge(qtr_final3, team_avg, how= 'left', on=['Qtr'])
qtr_final3.loc[:, "MCM Team Average"].fillna(0, inplace=True)

#CCR

#getting relevant columns for TA
ta_ccr1= ta_ccro[ta_ccro.columns[(0,8,9), ]].copy()
ta_ccr1= ta_ccr1.rename(columns={ta_ccr1.columns[0]:'Week'})
#ta_ccr1= ta_ccr1.drop(ta_ccr1[(ta_ccr1.loc[:, "Week"]!=w1)].index)

ta_ccr1=pd.merge(ta_ccr1, d2, on=['Week'])
ta_ccr1= ta_ccr1.groupby(['Qtr'])[['Correct Votes','Traditional CCR Sample']].sum()

ta_ccr1.reset_index(inplace=True)

ta_ccr1.loc[:, 'CCR Team Average']= (ta_ccr1.iloc[:, 1]/ ta_ccr1.iloc[:, 2])*100
ta_ccr1= ta_ccr1[['Qtr','CCR Team Average']]


qtr_final3= pd.merge(qtr_final3, ta_ccr1, how= 'left', on=['Qtr'])
qtr_final3.loc[:, "CCR Team Average"].fillna(0, inplace=True)

#EIM Escalations

#getting team average for EIM Esc
team_avg= eimo.groupby(['Week'])[['EIM Esc QA Sample','EIM Esc Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)

team_avg=pd.merge(team_avg, d2, how='left', on=['Week'])

team_avg= team_avg.groupby(['Qtr'])[['EIM Esc QA Sample','EIM Esc Total Errors']].sum()
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EIM Esc Team Average']= ((team_avg['EIM Esc QA Sample']- team_avg['EIM Esc Total Errors'])/team_avg['EIM Esc QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'EIM Esc Team Average']]

qtr_final3= pd.merge(qtr_final3, team_avg, how= 'left', on=['Qtr'])

qtr_final3.loc[:, "EIM Esc Team Average"].fillna(0, inplace=True)
#EIM FP

#getting team average for EIM FP
team_avg= eim_o.groupby(['Week'])[['EIM FP QA Sample','EIM FP Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)
team_avg=pd.merge(team_avg, d2, how='left', on=['Week'])

team_avg= team_avg.groupby(['Qtr'])[['EIM FP QA Sample','EIM FP Total Errors']].sum()
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EIM FP Team Average']= ((team_avg['EIM FP QA Sample']- team_avg['EIM FP Total Errors'])/team_avg['EIM FP QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'EIM FP Team Average']]

qtr_final3= pd.merge(qtr_final3, team_avg, how= 'left', on=['Qtr'])
qtr_final3.loc[:, "EIM FP Team Average"].fillna(0, inplace=True)

#Adhoc
adhoc2=adhoco[['Week', 'BO Team Average', 'Transparency Team Average']]

adhoc2=pd.merge(adhoc2, d2, on=['Week'], how='left')

adhoc2=adhoc2.groupby(['Qtr'])[['BO Team Average', 'Transparency Team Average']].mean()
adhoc2.reset_index(inplace=True)

adhoc2.loc[:, "BO Team Average"]=adhoc2.loc[:, "BO Team Average"]*100
adhoc2.loc[:, "Transparency Team Average"]=adhoc2.loc[:, "Transparency Team Average"]*100

qtr_final3=pd.merge(qtr_final3, adhoc2, how='left',on=['Qtr'])

qtr_final3.loc[:, "BO Team Average"].fillna(0, inplace=True)
qtr_final3.loc[:, "Transparency Team Average"].fillna(0, inplace=True)

#DCR

#getting team average for MCM

team_average= dcro.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= dcro.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Qtr'])[['QA Sample']].sum()
team_average= team_average.reset_index()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Qtr'])[['Total Errors']].sum()
team_average1= team_average1.reset_index()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Qtr'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'DCR Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'DCR Team Average']]


qtr_final3=pd.merge(qtr_final3, team_avg, how='left',on=['Qtr'])
qtr_final3.loc[:, "DCR Team Average"].fillna(0, inplace=True)


#Coverage

#getting team average for Coverage
team_average= coverageo.groupby(['Week'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= coverageo.groupby(['Week'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Qtr'])[['QA Sample']].sum()
team_average= team_average.reset_index()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Qtr'])[['Total Errors']].sum()
team_average1= team_average1.reset_index()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Qtr'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'Coverage Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'Coverage Team Average']]

qtr_final3=pd.merge(qtr_final3, team_avg, how='left',on=['Qtr'])
qtr_final3.loc[:, "Coverage Team Average"].fillna(0, inplace=True)


#EPCR
#getting team average for EPCR
team_average= epcr1.groupby(['Week'])[['EPCR QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= epcr1.groupby(['Week'])[['EPCR Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)
team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Qtr'])[['EPCR QA Sample']].sum()
team_average= team_average.reset_index()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Qtr'])[['EPCR Total Errors']].sum()
team_average1= team_average1.reset_index()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Qtr'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EPCR Team Average']= ((team_avg['EPCR QA Sample']- team_avg['EPCR Total Errors'])/team_avg['EPCR QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'EPCR Team Average']]

qtr_final3=pd.merge(qtr_final3, team_avg, how='left',on=['Qtr'])
qtr_final3.loc[:, "EPCR Team Average"].fillna(0, inplace=True)

#EACP appeals

#getting team average for EPCR
team_average= eacp_appeals.groupby(['Week'])[['EACP Appeals QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp_appeals.groupby(['Week'])[['EACP Appeals Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Qtr'])[['EACP Appeals QA Sample']].sum()
team_average= team_average.reset_index()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Qtr'])[['EACP Appeals Total Errors']].sum()
team_average1= team_average1.reset_index()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Qtr'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Appeals Team Average']= ((team_avg['EACP Appeals QA Sample']- team_avg['EACP Appeals Total Errors'])/team_avg['EACP Appeals QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'EACP Appeals Team Average']]

qtr_final3=pd.merge(qtr_final3, team_avg, how='left',on=['Qtr'])
qtr_final3.loc[:, "EACP Appeals Team Average"].fillna(0, inplace=True)

#EACP
#getting team average for EPCR
team_average= eacp1.groupby(['Week'])[['EACP QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp1.groupby(['Week'])[['EACP Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_average=pd.merge(team_average,d2,how='left', on=['Week'])
team_average1=pd.merge(team_average1,d2,how='left', on=['Week'])

team_average= team_average.groupby(['Qtr'])[['EACP QA Sample']].sum()
team_average= team_average.reset_index()
team_average= pd.DataFrame(team_average)
team_average1= team_average1.groupby(['Qtr'])[['EACP Total Errors']].sum()
team_average1= team_average1.reset_index()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['Qtr'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Team Average']= ((team_avg['EACP QA Sample']- team_avg['EACP Total Errors'])/team_avg['EACP QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['Qtr', 'EACP Team Average']]

qtr_final3=pd.merge(qtr_final3, team_avg, how='left',on=['Qtr'])
qtr_final3.loc[:, "EACP Team Average"].fillna(0, inplace=True)





qtr_final3=qtr_final3.reset_index()

qtr_final= pd.merge(qtr_final1, qtr_final3, on=['Qtr'])


#adding weighted average column to the final report

df=qtr_final.copy()

   
df.loc[:,'Associate Goal(Weighted Average)']= ((df.loc[:, "USN Team Average"]*df.loc[:, "USN Tagged Volume"])+
                                               (df.loc[:, "MCM Team Average"]*df.loc[:, "MCM Tagged Volume"])+
                                               (df.loc[:, "CCR Team Average"]*df.loc[:, "CCR Tagged Volume"])+
                                               (df.loc[:, "EIM Esc Team Average"]*df.loc[:, "EIM Esc Tagged Volume"])+
                                               (df.loc[:, "EIM FP Team Average"]*df.loc[:, "EIM FP Tagged Volume"])+
                                               (df.loc[:, "BO Team Average"]*df.loc[:, "BO Tagged Volume"])+
                                               (df.loc[:, "Transparency Team Average"]*df.loc[:, "Transparency Tagged Volume"])+
                                               (df.loc[:, "DCR Team Average"]*df.loc[:, "DCR Tagged Volume"])+
                                               (df.loc[:, "Coverage Team Average"]*df.loc[:, "Coverage Tagged Volume"])+
                                               (df.loc[:, "EPCR Team Average"]*df.loc[:, "EPCR Tagged Volume"])+
                                               (df.loc[:, "EACP Team Average"]*df.loc[:, "EACP Tagged Volume"])+
                                               (df.loc[:, "EACP Appeals Team Average"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                                                  df.loc[:, "MCM Tagged Volume"]+
                                                                                                                                  df.loc[:, "CCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                                                  df.loc[:, "BO Tagged Volume"]+
                                                                                                                                  df.loc[:, "Transparency Tagged Volume"]+
                                                                                                                                  df.loc[:, "DCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "Coverage Tagged Volume"]+
                                                                                                                                  df.loc[:, "EPCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Appeals Tagged Volume"])





df['Associate Goal(Weighted Average)'].fillna(0, inplace=True)
#adding deviation in USN
df.loc[:,'USN Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "USN Tagged Volume"], df.loc[:, "USN QA %"],
                                                                         (df.loc[:, "USN QA %"]-df.loc[:, "USN Team Average"]))]

df.loc[:,'MCM Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "MCM Tagged Volume"], df.loc[:, "MCM QA %"],
                                                                         (df.loc[:, "MCM QA %"]-df.loc[:, "MCM Team Average"]))]

df.loc[:,'CCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "CCR Tagged Volume"], df.loc[:, "CCR QA %"],
                                                                         (df.loc[:, "CCR QA %"]-df.loc[:, "CCR Team Average"]))]

df.loc[:,'EIM Esc Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM Esc Tagged Volume"], df.loc[:, "EIM Esc QA %"],
                                                                         (df.loc[:, "EIM Esc QA %"]-df.loc[:, "EIM Esc Team Average"]))]

df.loc[:,'EIM FP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM FP Tagged Volume"], df.loc[:, "EIM FP QA %"],
                                                                         (df.loc[:, "EIM FP QA %"]-df.loc[:, "EIM FP Team Average"]))]

df.loc[:,'BO Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "BO Tagged Volume"], df.loc[:, "BO QA %"],
                                                                         (df.loc[:, "BO QA %"]-df.loc[:, "BO Team Average"]))]

df.loc[:,'Transparency Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Transparency Tagged Volume"], df.loc[:, "Transparency QA %"],
                                                                         (df.loc[:, "Transparency QA %"]-df.loc[:, "Transparency Team Average"]))]

df.loc[:,'DCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "DCR Tagged Volume"], df.loc[:, "DCR QA %"],
                                                                         (df.loc[:, "DCR QA %"]-df.loc[:, "DCR Team Average"]))]

df.loc[:,'Coverage Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Coverage Tagged Volume"], df.loc[:, "Coverage QA %"],
                                                                         (df.loc[:, "Coverage QA %"]-df.loc[:, "Coverage Team Average"]))]

df.loc[:,'EPCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EPCR Tagged Volume"], df.loc[:, "EPCR QA %"],
                                                                         (df.loc[:, "EPCR QA %"]-df.loc[:, "EPCR Team Average"]))]

df.loc[:,'EACP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Tagged Volume"], df.loc[:, "EACP QA %"],
                                                                         (df.loc[:, "EACP QA %"]-df.loc[:, "EACP Team Average"]))]

df.loc[:,'EACP Appeals Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Appeals Tagged Volume"], df.loc[:, "EACP Appeals QA %"],
                                                                         (df.loc[:, "EACP Appeals QA %"]-df.loc[:, "EACP Appeals Team Average"]))]






#net deviation column to the final report
df.loc[:,'Net Deviation']= ((df.loc[:, "USN Deviation"]*df.loc[:, "USN Tagged Volume"])+
                            (df.loc[:, "MCM Deviation"]*df.loc[:, "MCM Tagged Volume"])+
                            (df.loc[:, "CCR Deviation"]*df.loc[:, "CCR Tagged Volume"])+
                            (df.loc[:, "EIM Esc Deviation"]*df.loc[:, "EIM Esc Tagged Volume"])+
                            (df.loc[:, "EIM FP Deviation"]*df.loc[:, "EIM FP Tagged Volume"])+
                            (df.loc[:, "BO Deviation"]*df.loc[:, "BO Tagged Volume"])+
                            (df.loc[:, "Transparency Deviation"]*df.loc[:, "Transparency Tagged Volume"])+
                            (df.loc[:, "DCR Deviation"]*df.loc[:, "DCR Tagged Volume"])+
                            (df.loc[:, "Coverage Deviation"]*df.loc[:, "Coverage Tagged Volume"])+
                            (df.loc[:, "EPCR Deviation"]*df.loc[:, "EPCR Tagged Volume"])+
                            (df.loc[:, "EACP Deviation"]*df.loc[:, "EACP Tagged Volume"])+
                            (df.loc[:, "EACP Appeals Deviation"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                            df.loc[:, "MCM Tagged Volume"]+
                                                                                                            df.loc[:, "CCR Tagged Volume"]+
                                                                                                            df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                            df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                            df.loc[:, "BO Tagged Volume"]+
                                                                                                            df.loc[:, "Transparency Tagged Volume"]+
                                                                                                            df.loc[:, "DCR Tagged Volume"]+
                                                                                                            df.loc[:, "Coverage Tagged Volume"]+
                                                                                                            df.loc[:, "EPCR Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Appeals Tagged Volume"])



df['Net Deviation'].fillna(0, inplace=True)
#Final QA% (Weighed Average)
df.loc[:,'Final QA%(Weighted Average)']= [x if y==0 else z for x,y,z in zip(df.loc[:,'Associate Goal(Weighted Average)'],df.loc[:, "Net Deviation"],(df.loc[:,'Associate Goal(Weighted Average)']+ df.loc[:,'Net Deviation']))]

#converting NaN to 0

df.replace([np.inf, -np.inf], np.nan, inplace=True)
df= df.fillna(0)

df['Final QA%(Weighted Average)'].fillna(0, inplace=True)



df.loc[:, "Year"]=year_data

#converting NaN to 0
df.replace([np.inf, -np.inf], np.nan, inplace=True)
df= df.fillna(0)



#adding name column from login sheet


qtr_final= pd.merge(df, d1, how='left', on=['Login'])
qtr_final.sort_values(by=['Login'], inplace=True)

qtr_final.loc[:, "Total QA Sample"]=[np.nan if x=="Yes" else y for x,y in zip(qtr_final.loc[:, "Shared Associate"],(qtr_final.loc[:, "USN QA Sample"]+qtr_final.loc[:, "MCM QA Sample"]+qtr_final.loc[:, "CCR QA Sample"]+
                                                                                                                    qtr_final.loc[:, "EIM Esc QA Sample"]+qtr_final.loc[:, "BO QA Sample Audits"]+
                                                                                                                    qtr_final.loc[:, "Transparency QA Sample Audits"]+qtr_final.loc[:, "DCR QA Sample"]+qtr_final.loc[:, "EPCR QA Sample"]+
                                                                                                                    qtr_final.loc[:, "EACP QA Sample"]+qtr_final.loc[:, "EACP Appeals QA Sample"]+qtr_final.loc[:, "Coverage QA Sample"]+qtr_final.loc[:, "EIM FP QA Sample"]))]

qtr_final.loc[:, "Total Errors"]= [np.nan if x=="Yes" else y for x,y in zip(qtr_final.loc[:, "Shared Associate"],(qtr_final.loc[:, "USN Total Errors"]+qtr_final.loc[:, "MCM Total Errors"]+
                                                                                                                  qtr_final.loc[:, "CCR Total Errors"]+qtr_final.loc[:, "EIM Esc Total Errors"]+
                                                                                                                  qtr_final.loc[:, "BO Total Errors"]+
                                                                                                                  qtr_final.loc[:, "Transparency Total Errors"]+qtr_final.loc[:, "EIM FP Total Errors"]+
                                                                                                                  qtr_final.loc[:, "DCR Total Errors"]+qtr_final.loc[:, "EPCR Total Errors"]+
                                                                                                                  qtr_final.loc[:, "EACP Total Errors"]+qtr_final.loc[:, "EACP Appeals Total Errors"]+
                                                                                                                  qtr_final.loc[:, "Coverage Total Errors"]))]


qtr_final.loc[:, "QA % (Non Shared Associates)"]=[np.nan if x=='Yes' else y for x,y in zip (qtr_final.loc[:, "Shared Associate"],
                                                                                            ((qtr_final.loc[:, "Total QA Sample"]-qtr_final.loc[:, "Total Errors"])/qtr_final.loc[:,"Total QA Sample"]))]
        
qtr_final.loc[:, "QA % (Non Shared Associates)"]= qtr_final.loc[:, 'QA % (Non Shared Associates)']*100

qtr_final.loc[:, "Final QA%(Weighted Average)"]=[np.nan if x=='No' else y for x,y in zip (qtr_final.loc[:, "Shared Associate"],qtr_final.loc[:, "Final QA%(Weighted Average)"])]
 
qtr_final.loc[:, "Weighted Productivity %"]= [0 if x==0 else y for x,y in zip(qtr_final.loc[:, "Actual Cumputed Hours"],(qtr_final.loc[:, "Weighted Productive Sum"]/qtr_final.loc[:, "Actual Cumputed Hours"]))]
qtr_final.loc[:, "Weighted Productivity %"]=qtr_final.loc[:, "Weighted Productivity %"]*100
qtr_final.loc[:, "Year"]= year_data

qtr_final.sort_values(by=['Qtr', 'Login'], ascending=True, inplace=True)
#getting columns position as per the report

df7= qtr_final[['Name','Login','Program','Manager','Shared Associate','Year', 'Qtr',
                'USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
                'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
                'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
                'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
                'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
                'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
                'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
                'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
                'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
                'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
                'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
                'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
                'Associate Goal(Weighted Average)', 'USN Deviation',
                'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
                'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
                'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation',
                'Final QA%(Weighted Average)','Total QA Sample', 'Total Errors', 'QA % (Non Shared Associates)',
                'Total','Actual Cumputed Hours','Actual Cumputed NPT',
                'Leave Hours','Weighted Productive Sum','Weighted Productivity %']]


cols=['Total','Actual Cumputed Hours','Actual Cumputed NPT',
      'Leave Hours','Weighted Productive Sum','Weighted Productivity %']

df7[cols].replace(to_replace= 0, value= np.nan, inplace= True)

df7= df7.round({"Final QA%(Weighted Average)":2, "Weighted Productivity %":2, "QA % (Non Shared Associates) %":2})
        



    
quarterly_consolidated= df7[['Name','Login','Program','Manager','Shared Associate','Year','Qtr',
                              'Final QA%(Weighted Average)','Total QA Sample', 
                              'Total Errors', 'QA % (Non Shared Associates)', 'Weighted Productivity %']].copy()
    
#YTD report
ytd1=data.copy()

#creating a new dataframe in the same sheet
ytd2= ytd1.groupby(['Login'])[['USN Tagged Volume','MCM Tagged Volume', 'CCR Tagged Volume', 'EIM Esc Tagged Volume',
                               'EIM FP Tagged Volume','BO Tagged Volume',
                               'EPCR Tagged Volume','EACP Tagged Volume','EACP Appeals Tagged Volume',
                               'DCR Tagged Volume', 'Coverage Tagged Volume',
                               'Transparency Tagged Volume','USN Total Errors','USN QA Sample',
                               'MCM Total Errors','MCM QA Sample','CCR Total Errors','CCR QA Sample',
                               'EIM Esc Total Errors', 'EIM Esc QA Sample','EIM FP Total Errors', 'EIM FP QA Sample',
                               'BO QA Sample Audits','BO Total Errors',
                               'Transparency QA Sample Audits', 'Transparency Total Errors',
                               'DCR QA Sample', 'DCR Total Errors','Coverage QA Sample', 'Coverage Total Errors',
                               'EPCR QA Sample', 'EPCR Total Errors','EACP QA Sample', 'EACP Total Errors',
                               'EACP Appeals QA Sample', 'EACP Appeals Total Errors','Total QA Sample', 'Total Errors',
                               'Total','Actual Cumputed Hours','Actual Cumputed NPT',
                               'Leave Hours','Weighted Productive Sum']].sum()
ytd2= pd.DataFrame(ytd2)
ytd2= ytd2.reset_index()


ytd2.loc[:,'USN QA %']= (ytd2.iloc[:,10]
                          -ytd2.iloc[:,9])/ytd2.iloc[:,10]
ytd2.loc[:,'USN QA %']= (ytd2.loc[:, "USN QA Sample"]
                          -ytd2.loc[:,"USN Total Errors"])/ytd2.loc[:,"USN QA Sample"]
ytd2.loc[:,'MCM QA %']= (ytd2.loc[:, "MCM QA Sample"]
                          -ytd2.loc[:,"MCM Total Errors"])/ytd2.loc[:,"MCM QA Sample"]
ytd2.loc[:,'CCR QA %']= (ytd2.loc[:, "CCR QA Sample"]
                          -ytd2.loc[:,"CCR Total Errors"])/ytd2.loc[:,"CCR QA Sample"]
ytd2.loc[:,'EIM Esc QA %']= (ytd2.loc[:, "EIM Esc QA Sample"]
                          -ytd2.loc[:,"EIM Esc Total Errors"])/ytd2.loc[:,"EIM Esc QA Sample"]
ytd2.loc[:,'EIM FP QA %']= (ytd2.loc[:, "EIM FP QA Sample"]
                          -ytd2.loc[:,"EIM FP Total Errors"])/ytd2.loc[:,"EIM FP QA Sample"]
ytd2.loc[:,'BO QA %']= (ytd2.loc[:, "BO QA Sample Audits"]
                          -ytd2.loc[:,"BO Total Errors"])/ytd2.loc[:,"BO QA Sample Audits"]
ytd2.loc[:,'Transparency QA %']= (ytd2.loc[:, "Transparency QA Sample Audits"]
                          -ytd2.loc[:,"Transparency Total Errors"])/ytd2.loc[:,"Transparency QA Sample Audits"]
ytd2.loc[:,'EPCR QA %']= (ytd2.loc[:, "EPCR QA Sample"]
                          -ytd2.loc[:,"EPCR Total Errors"])/ytd2.loc[:,"EPCR QA Sample"]
ytd2.loc[:,'EACP QA %']= (ytd2.loc[:, "EACP QA Sample"]
                          -ytd2.loc[:,"EACP Total Errors"])/ytd2.loc[:,"EACP QA Sample"]
ytd2.loc[:,'EACP Appeals QA %']= (ytd2.loc[:, "EACP Appeals QA Sample"]
                                   -ytd2.loc[:,"EACP Appeals Total Errors"])/ytd2.loc[:,"EACP Appeals QA Sample"]
ytd2.loc[:,'DCR QA %']= (ytd2.loc[:, "DCR QA Sample"]
                          -ytd2.loc[:,"DCR Total Errors"])/ytd2.loc[:,"DCR QA Sample"]
ytd2.loc[:,'Coverage QA %']= (ytd2.loc[:, "Coverage QA Sample"]
                          -ytd2.loc[:,"Coverage Total Errors"])/ytd2.loc[:,"Coverage QA Sample"]


#converting accuracy columns into percentage by multiplying by 100
ytd2['MCM QA %']= ytd2['MCM QA %']*100
ytd2['USN QA %']= ytd2['USN QA %']*100
ytd2['CCR QA %']= ytd2['CCR QA %']*100
ytd2['EIM Esc QA %']= ytd2['EIM Esc QA %']*100
ytd2['EIM FP QA %']= ytd2['EIM FP QA %']*100
ytd2['BO QA %']= ytd2['BO QA %']*100
ytd2['Transparency QA %']= ytd2['Transparency QA %']*100
ytd2['EPCR QA %']= ytd2['EPCR QA %']*100
ytd2['EACP QA %']= ytd2['EACP QA %']*100
ytd2['EACP Appeals QA %']= ytd2['EACP Appeals QA %']*100
ytd2['DCR QA %']= ytd2['DCR QA %']*100
ytd2['Coverage QA %']= ytd2['Coverage QA %']*100


ytd2.sort_values(by=['Login'], inplace=True)



#YTD team average


#USN

ta_usno= ta_usno.rename(columns={ta_usno.columns[0]:'Week'})
#ta_usno= ta_usno.drop(ta_usno[(ta_usno.loc[:, "Week"]!=w1)].index)

ta_usn1= ta_usno[ta_usno.columns[(0,7,9), ]].copy()
ta_usn1.loc[:, "temp"]=1

ta_usn1= ta_usn1.groupby(['temp'])[['TMM USN Sample','Correct Votes']].sum()
ta_usn1.reset_index(inplace=True)

ta_usn1.loc[:, 'USN Team Average']= (ta_usn1.iloc[:, 2]/ta_usn1.iloc[:, 1])*100

ta_usn1= ta_usn1[['temp','USN Team Average']]



ytd2.loc[:, "temp"]=1

ytd2= pd.merge(ytd2, ta_usn1, how= 'left', on=['temp'])

ytd2.loc[:, "USN Team Average"].fillna(0, inplace=True)
#MCM


#getting team average for MCM
mcm_ta=mcmo.copy()
mcm_ta.loc[:, "temp"]=1
team_average= mcm_ta.groupby(['temp'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= mcm_ta.groupby(['temp'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg_mcm= pd.merge(team_average, team_average1, on=['temp'])
team_avg_mcm= pd.DataFrame(team_avg_mcm)
team_avg=team_avg_mcm.copy()

team_avg.loc[:,'MCM Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'MCM Team Average']] 

ytd2= pd.merge(ytd2, team_avg, how= 'left', on=['temp'])
ytd2.loc[:, "MCM Team Average"].fillna(0, inplace=True)

#CCR

#getting relevant columns for TA

ta_ccr1= ta_ccro[ta_ccro.columns[(0,8,9), ]].copy()
ta_ccr1.loc[:, "temp"]=1
ta_ccr1= ta_ccr1.rename(columns={ta_ccr1.columns[0]:'Week'})
#ta_ccr1= ta_ccr1.drop(ta_ccr1[(ta_ccr1.loc[:, "Week"]!=w1)].index)

ta_ccr1= ta_ccr1.groupby(['temp'])[['Correct Votes','Traditional CCR Sample']].sum()
ta_ccr1.reset_index(inplace=True)
ta_ccr1.loc[:, 'CCR Team Average']= (ta_ccr1.iloc[:, 1]/ ta_ccr1.iloc[:, 2])*100
ta_ccr1= ta_ccr1[['temp','CCR Team Average']]


ytd2= pd.merge(ytd2, ta_ccr1, how= 'left', on=['temp'])
ytd2.loc[:, "CCR Team Average"].fillna(0, inplace=True)

#EIM Escalations

#getting team average for EIM Esc
eim_ta=eimo.copy()
eim_ta.loc[:, "temp"]=1
team_avg= eim_ta.groupby(['temp'])[['EIM Esc QA Sample','EIM Esc Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)



team_avg.loc[:,'EIM Esc Team Average']= ((team_avg['EIM Esc QA Sample']- team_avg['EIM Esc Total Errors'])/team_avg['EIM Esc QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'EIM Esc Team Average']]

ytd2= pd.merge(ytd2, team_avg, how= 'left', on=['temp'])

ytd2.loc[:, "EIM Esc Team Average"].fillna(0, inplace=True)
#EIM FP

#getting team average for EIM FP
eimfp_ta=eim_o.copy()
eimfp_ta.loc[:, "temp"]=1
team_avg= eimfp_ta.groupby(['temp'])[['EIM FP QA Sample','EIM FP Total Errors']].sum()
team_avg= pd.DataFrame(team_avg)


team_avg.loc[:,'EIM FP Team Average']= ((team_avg['EIM FP QA Sample']- team_avg['EIM FP Total Errors'])/team_avg['EIM FP QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'EIM FP Team Average']]

ytd2= pd.merge(ytd2, team_avg, how= 'left', on=['temp'])
ytd2.loc[:, "EIM FP Team Average"].fillna(0, inplace=True)

#Adhoc
adhoc2=adhoco[['BO Team Average', 'Transparency Team Average']]
adhoc2.loc[:, "temp"]=1

adhoc2=adhoc2.groupby(['temp'])[['BO Team Average', 'Transparency Team Average']].mean()
adhoc2.reset_index(inplace=True)


adhoc2=adhoc2[['temp','BO Team Average', 'Transparency Team Average']]

adhoc2.loc[:, "BO Team Average"]=adhoc2.loc[:, "BO Team Average"]*100
adhoc2.loc[:, "Transparency Team Average"]=adhoc2.loc[:, "Transparency Team Average"]*100

ytd2=pd.merge(ytd2, adhoc2, how='left',on=['temp'])

ytd2.loc[:, "BO Team Average"].fillna(0, inplace=True)
ytd2.loc[:, "Transparency Team Average"].fillna(0, inplace=True)

#DCR

#getting team average for MCM
dcr_ta=dcro.copy()
dcr_ta.loc[:, "temp"]=1
team_average= dcr_ta.groupby(['temp'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= dcr_ta.groupby(['temp'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)


team_avg= pd.merge(team_average, team_average1, on=['temp'])
team_avg= pd.DataFrame(team_avg)



team_avg.loc[:,'DCR Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'DCR Team Average']]


ytd2=pd.merge(ytd2, team_avg, how='left',on=['temp'])
ytd2.loc[:, "DCR Team Average"].fillna(0, inplace=True)


#Coverage

#getting team average for Coverage
coverage_ta=coverageo.copy()
coverage_ta.loc[:, "temp"]=1
team_average= coverage_ta.groupby(['temp'])[['QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= coverage_ta.groupby(['temp'])[['Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)


team_avg= pd.merge(team_average, team_average1, on=['temp'])
team_avg= pd.DataFrame(team_avg)



team_avg.loc[:,'Coverage Team Average']= ((team_avg['QA Sample']- team_avg['Total Errors'])/team_avg['QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'Coverage Team Average']]

ytd2=pd.merge(ytd2, team_avg, how='left',on=['temp'])
ytd2.loc[:, "Coverage Team Average"].fillna(0, inplace=True)


#EPCR
#getting team average for EPCR
epcr_ta=epcr1.copy()
epcr_ta.loc[:, "temp"]=1
team_average= epcr_ta.groupby(['temp'])[['EPCR QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= epcr_ta.groupby(['temp'])[['EPCR Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)


team_avg= pd.merge(team_average, team_average1, on=['temp'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EPCR Team Average']= ((team_avg['EPCR QA Sample']- team_avg['EPCR Total Errors'])/team_avg['EPCR QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'EPCR Team Average']]

ytd2=pd.merge(ytd2, team_avg, how='left',on=['temp'])
ytd2.loc[:, "EPCR Team Average"].fillna(0, inplace=True)

#EACP appeals

#getting team average for EPCR
eacp_appeals_ta=eacp_appeals.copy()
eacp_appeals_ta.loc[:, "temp"]=1
team_average= eacp_appeals_ta.groupby(['temp'])[['EACP Appeals QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp_appeals_ta.groupby(['temp'])[['EACP Appeals Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)


team_avg= pd.merge(team_average, team_average1, on=['temp'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Appeals Team Average']= ((team_avg['EACP Appeals QA Sample']- team_avg['EACP Appeals Total Errors'])/team_avg['EACP Appeals QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'EACP Appeals Team Average']]

ytd2=pd.merge(ytd2, team_avg, how='left',on=['temp'])
ytd2.loc[:, "EACP Appeals Team Average"].fillna(0, inplace=True)

#EACP
#getting team average for EPCR
eacp_ta=eacp1.copy()
eacp_ta.loc[:, "temp"]=1
team_average= eacp_ta.groupby(['temp'])[['EACP QA Sample']].sum()
team_average= pd.DataFrame(team_average)
team_average1= eacp_ta.groupby(['temp'])[['EACP Total Errors']].sum()
team_average1= pd.DataFrame(team_average1)

team_avg= pd.merge(team_average, team_average1, on=['temp'])
team_avg= pd.DataFrame(team_avg)

team_avg.loc[:,'EACP Team Average']= ((team_avg['EACP QA Sample']- team_avg['EACP Total Errors'])/team_avg['EACP QA Sample'])*100
team_avg= team_avg.reset_index()
team_avg= pd.DataFrame(team_avg)
team_avg= team_avg[['temp', 'EACP Team Average']]

ytd2=pd.merge(ytd2, team_avg, how='left',on=['temp'])
ytd2.loc[:, "EACP Team Average"].fillna(0, inplace=True)


#adding weighted average column to the final report

df=ytd2.copy()

df.loc[:,'Associate Goal(Weighted Average)']= ((df.loc[:, "USN Team Average"]*df.loc[:, "USN Tagged Volume"])+
                                               (df.loc[:, "MCM Team Average"]*df.loc[:, "MCM Tagged Volume"])+
                                               (df.loc[:, "CCR Team Average"]*df.loc[:, "CCR Tagged Volume"])+
                                               (df.loc[:, "EIM Esc Team Average"]*df.loc[:, "EIM Esc Tagged Volume"])+
                                               (df.loc[:, "EIM FP Team Average"]*df.loc[:, "EIM FP Tagged Volume"])+
                                               (df.loc[:, "BO Team Average"]*df.loc[:, "BO Tagged Volume"])+
                                               (df.loc[:, "Transparency Team Average"]*df.loc[:, "Transparency Tagged Volume"])+
                                               (df.loc[:, "DCR Team Average"]*df.loc[:, "DCR Tagged Volume"])+
                                               (df.loc[:, "Coverage Team Average"]*df.loc[:, "Coverage Tagged Volume"])+
                                               (df.loc[:, "EPCR Team Average"]*df.loc[:, "EPCR Tagged Volume"])+
                                               (df.loc[:, "EACP Team Average"]*df.loc[:, "EACP Tagged Volume"])+
                                               (df.loc[:, "EACP Appeals Team Average"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                                                  df.loc[:, "MCM Tagged Volume"]+
                                                                                                                                  df.loc[:, "CCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                                                  df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                                                  df.loc[:, "BO Tagged Volume"]+
                                                                                                                                  df.loc[:, "Transparency Tagged Volume"]+
                                                                                                                                  df.loc[:, "DCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "Coverage Tagged Volume"]+
                                                                                                                                  df.loc[:, "EPCR Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Tagged Volume"]+
                                                                                                                                  df.loc[:, "EACP Appeals Tagged Volume"])





df['Associate Goal(Weighted Average)'].fillna(0, inplace=True)
#adding deviation in USN
df.loc[:,'USN Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "USN Tagged Volume"], df.loc[:, "USN QA %"],
                                                                         (df.loc[:, "USN QA %"]-df.loc[:, "USN Team Average"]))]

df.loc[:,'MCM Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "MCM Tagged Volume"], df.loc[:, "MCM QA %"],
                                                                         (df.loc[:, "MCM QA %"]-df.loc[:, "MCM Team Average"]))]

df.loc[:,'CCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "CCR Tagged Volume"], df.loc[:, "CCR QA %"],
                                                                         (df.loc[:, "CCR QA %"]-df.loc[:, "CCR Team Average"]))]

df.loc[:,'EIM Esc Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM Esc Tagged Volume"], df.loc[:, "EIM Esc QA %"],
                                                                         (df.loc[:, "EIM Esc QA %"]-df.loc[:, "EIM Esc Team Average"]))]

df.loc[:,'EIM FP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EIM FP Tagged Volume"], df.loc[:, "EIM FP QA %"],
                                                                         (df.loc[:, "EIM FP QA %"]-df.loc[:, "EIM FP Team Average"]))]

df.loc[:,'BO Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "BO Tagged Volume"], df.loc[:, "BO QA %"],
                                                                         (df.loc[:, "BO QA %"]-df.loc[:, "BO Team Average"]))]

df.loc[:,'Transparency Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Transparency Tagged Volume"], df.loc[:, "Transparency QA %"],
                                                                         (df.loc[:, "Transparency QA %"]-df.loc[:, "Transparency Team Average"]))]

df.loc[:,'DCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "DCR Tagged Volume"], df.loc[:, "DCR QA %"],
                                                                         (df.loc[:, "DCR QA %"]-df.loc[:, "DCR Team Average"]))]

df.loc[:,'Coverage Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "Coverage Tagged Volume"], df.loc[:, "Coverage QA %"],
                                                                         (df.loc[:, "Coverage QA %"]-df.loc[:, "Coverage Team Average"]))]

df.loc[:,'EPCR Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EPCR Tagged Volume"], df.loc[:, "EPCR QA %"],
                                                                         (df.loc[:, "EPCR QA %"]-df.loc[:, "EPCR Team Average"]))]

df.loc[:,'EACP Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Tagged Volume"], df.loc[:, "EACP QA %"],
                                                                         (df.loc[:, "EACP QA %"]-df.loc[:, "EACP Team Average"]))]

df.loc[:,'EACP Appeals Deviation'] = [0 if x == 0 or z==0 else y for x,z,y in zip(df.loc[:, "EACP Appeals Tagged Volume"], df.loc[:, "EACP Appeals QA %"],
                                                                         (df.loc[:, "EACP Appeals QA %"]-df.loc[:, "EACP Appeals Team Average"]))]






#net deviation column to the final report
df.loc[:,'Net Deviation']= ((df.loc[:, "USN Deviation"]*df.loc[:, "USN Tagged Volume"])+
                            (df.loc[:, "MCM Deviation"]*df.loc[:, "MCM Tagged Volume"])+
                            (df.loc[:, "CCR Deviation"]*df.loc[:, "CCR Tagged Volume"])+
                            (df.loc[:, "EIM Esc Deviation"]*df.loc[:, "EIM Esc Tagged Volume"])+
                            (df.loc[:, "EIM FP Deviation"]*df.loc[:, "EIM FP Tagged Volume"])+
                            (df.loc[:, "BO Deviation"]*df.loc[:, "BO Tagged Volume"])+
                            (df.loc[:, "Transparency Deviation"]*df.loc[:, "Transparency Tagged Volume"])+
                            (df.loc[:, "DCR Deviation"]*df.loc[:, "DCR Tagged Volume"])+
                            (df.loc[:, "Coverage Deviation"]*df.loc[:, "Coverage Tagged Volume"])+
                            (df.loc[:, "EPCR Deviation"]*df.loc[:, "EPCR Tagged Volume"])+
                            (df.loc[:, "EACP Deviation"]*df.loc[:, "EACP Tagged Volume"])+
                            (df.loc[:, "EACP Appeals Deviation"]*df.loc[:, "EACP Appeals Tagged Volume"]))/(df.loc[:, "USN Tagged Volume"]+
                                                                                                            df.loc[:, "MCM Tagged Volume"]+
                                                                                                            df.loc[:, "CCR Tagged Volume"]+
                                                                                                            df.loc[:, "EIM Esc Tagged Volume"]+
                                                                                                            df.loc[:, "EIM FP Tagged Volume"]+
                                                                                                            df.loc[:, "BO Tagged Volume"]+
                                                                                                            df.loc[:, "Transparency Tagged Volume"]+
                                                                                                            df.loc[:, "DCR Tagged Volume"]+
                                                                                                            df.loc[:, "Coverage Tagged Volume"]+
                                                                                                            df.loc[:, "EPCR Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Tagged Volume"]+
                                                                                                            df.loc[:, "EACP Appeals Tagged Volume"])



df['Net Deviation'].fillna(0, inplace=True)
#Final QA% (Weighed Average)
df.loc[:,'Final QA%(Weighted Average)']= [x if y==0 else z for x,y,z in zip(df.loc[:,'Associate Goal(Weighted Average)'],df.loc[:, "Net Deviation"],(df.loc[:,'Associate Goal(Weighted Average)']+ df.loc[:,'Net Deviation']))]

#converting NaN to 0

df.replace([np.inf, -np.inf], np.nan, inplace=True)


df['Final QA%(Weighted Average)'].fillna(0, inplace=True)



df.loc[:, "Year"]=year_data
   

#converting NaN to 0
df.replace([np.inf, -np.inf], np.nan, inplace=True)


#adding name column from the login sheet


ytd2= pd.merge(df, d1, how='left', on=['Login'])

ytd2.loc[:, "Total QA Sample"]=[np.nan if x=="Yes" else y for x,y in zip(ytd2.loc[:, "Shared Associate"],(ytd2.loc[:, "USN QA Sample"]+ytd2.loc[:, "MCM QA Sample"]+ytd2.loc[:, "CCR QA Sample"]+
                                                                                                          ytd2.loc[:, "EIM Esc QA Sample"]+ytd2.loc[:, "BO QA Sample Audits"]+
                                                                                                          ytd2.loc[:, "Transparency QA Sample Audits"]+ytd2.loc[:, "DCR QA Sample"]+ytd2.loc[:, "EPCR QA Sample"]+
                                                                                                          ytd2.loc[:, "EACP QA Sample"]+ytd2.loc[:, "EACP Appeals QA Sample"]+ytd2.loc[:, "Coverage QA Sample"]+ytd2.loc[:, "EIM FP QA Sample"]))]

ytd2.loc[:, "Total Errors"]= [np.nan if x=="Yes" else y for x,y in zip(ytd2.loc[:, "Shared Associate"],(ytd2.loc[:, "USN Total Errors"]+ytd2.loc[:, "MCM Total Errors"]+
                                                                                                        ytd2.loc[:, "CCR Total Errors"]+ytd2.loc[:, "EIM Esc Total Errors"]+
                                                                                                        ytd2.loc[:, "BO Total Errors"]+
                                                                                                        ytd2.loc[:, "Transparency Total Errors"]+ytd2.loc[:, "EIM FP Total Errors"]+
                                                                                                        ytd2.loc[:, "DCR Total Errors"]+ytd2.loc[:, "EPCR Total Errors"]+
                                                                                                        ytd2.loc[:, "EACP Total Errors"]+ytd2.loc[:, "EACP Appeals Total Errors"]+
                                                                                                        ytd2.loc[:, "Coverage Total Errors"]))]


ytd2.loc[:, "QA % (Non Shared Associates)"]=[np.nan if x=='Yes' else y for x,y in zip (ytd2.loc[:, "Shared Associate"],
                                                                                   ((ytd2.loc[:, "Total QA Sample"]-ytd2.loc[:, "Total Errors"])/ytd2.loc[:,"Total QA Sample"]))]
        
ytd2.loc[:, "QA % (Non Shared Associates)"]= ytd2.loc[:, 'QA % (Non Shared Associates)']*100

ytd2.loc[:, "Final QA%(Weighted Average)"]=[np.nan if x=='No' else y for x,y in zip (ytd2.loc[:, "Shared Associate"],ytd2.loc[:, "Final QA%(Weighted Average)"])]
        
ytd2.loc[:, "Weighted Productivity %"]= [0 if x==0 else y for x,y in zip(ytd2.loc[:, "Actual Cumputed Hours"],(ytd2.loc[:, "Weighted Productive Sum"]/ytd2.loc[:, "Actual Cumputed Hours"]))]
ytd2.loc[:, "Weighted Productivity %"]=ytd2.loc[:, "Weighted Productivity %"]*100
#getting columns position as per the report

ytd_final= ytd2[['Name','Login','Program','Manager','Shared Associate','Year',
                 'USN Tagged Volume','USN QA Sample', 'USN Total Errors','USN QA %', 'USN Team Average',
                 'MCM Tagged Volume','MCM QA Sample','MCM Total Errors', 'MCM QA %', 'MCM Team Average',
                 'CCR Tagged Volume','CCR QA Sample', 'CCR Total Errors', 'CCR QA %', 'CCR Team Average',
                 'EIM Esc Tagged Volume','EIM Esc QA Sample','EIM Esc Total Errors','EIM Esc QA %', 'EIM Esc Team Average',
                 'EIM FP Tagged Volume','EIM FP QA Sample','EIM FP Total Errors','EIM FP QA %', 'EIM FP Team Average',
                 'BO Tagged Volume','BO QA Sample Audits','BO Total Errors', 'BO QA %', 'BO Team Average',
                 'Transparency Tagged Volume','Transparency QA Sample Audits', 'Transparency Total Errors', 'Transparency QA %', 'Transparency Team Average',
                 'DCR Tagged Volume','DCR QA Sample', 'DCR Total Errors', 'DCR QA %', 'DCR Team Average',
                 'EPCR Tagged Volume','EPCR QA Sample', 'EPCR Total Errors', 'EPCR QA %', 'EPCR Team Average',
                 'EACP Tagged Volume','EACP QA Sample', 'EACP Total Errors', 'EACP QA %', 'EACP Team Average',
                 'EACP Appeals Tagged Volume','EACP Appeals QA Sample', 'EACP Appeals Total Errors', 'EACP Appeals QA %', 'EACP Appeals Team Average',
                 'Coverage Tagged Volume', 'Coverage QA Sample', 'Coverage Total Errors', 'Coverage QA %', 'Coverage Team Average',
                 'Associate Goal(Weighted Average)', 'USN Deviation',
                 'MCM Deviation','CCR Deviation','EIM Esc Deviation','EIM FP Deviation','BO Deviation',
                 'Transparency Deviation','DCR Deviation', 'EPCR Deviation',
                 'EACP Deviation', 'EACP Appeals Deviation','Coverage Deviation','Net Deviation',
                 'Final QA%(Weighted Average)','Total QA Sample', 'Total Errors', 'QA % (Non Shared Associates)',
                 'Total','Actual Cumputed Hours','Actual Cumputed NPT',
                 'Leave Hours','Weighted Productive Sum','Weighted Productivity %']].copy()

cols=['Weighted Productivity %']

ytd_final[cols].replace(to_replace= 0, value= np.nan, inplace= True)

ytd_final= ytd_final.round({"Final QA%(Weighted Average)":2, "Weighted Productivity %":2, "QA % (Non Shared Associates) %":2})
    

#YTD consolidated
consolidated_ytd=ytd_final[['Name','Login','Program','Manager','Shared Associate','Year',
                            'Final QA%(Weighted Average)','Total QA Sample', 
                            'Total Errors', 'QA % (Non Shared Associates)', 'Weighted Productivity %']]

data.replace(to_replace= 0, value= np.nan, inplace= True)
df5.replace(to_replace= 0, value= np.nan, inplace= True)
df7.replace(to_replace= 0, value= np.nan, inplace= True)
ytd_final.replace(to_replace= 0, value= np.nan, inplace= True)

data_final=data.copy()
df_5=df5.copy()
df_7=df7.copy()
df_8=ytd_final.copy()
writer = pd.ExcelWriter(path1+"/Final_Report"+"/Overall_Productivity_Quality_Report_Till_Week"+str(w1)+".xlsx", engine='xlsxwriter')
data_final.to_excel(writer, sheet_name='Weekly Report', index=False)  # Default position, cell A1.
df_5.to_excel(writer, sheet_name='Monthly Report', index=False)
df_7.to_excel(writer, sheet_name='Quarterly Report', index=False)
df_8.to_excel(writer, sheet_name='YTD', index=False)
writer.save()

consolidated_weekly.replace(to_replace= 0, value= np.nan, inplace= True)
monthly_consolidated.replace(to_replace= 0, value= np.nan, inplace= True)
quarterly_consolidated.replace(to_replace= 0, value= np.nan, inplace= True)
consolidated_ytd.replace(to_replace= 0, value= np.nan, inplace= True)

weekly_1=consolidated_weekly.copy()
monthly_1=monthly_consolidated.copy()
quarterly_1=quarterly_consolidated.copy()
df_9=consolidated_ytd.copy()
writer2= pd.ExcelWriter(path1+"/Final_Report"+"/Overall_Consolidated_Report_Till_Week_"+str(w1)+".xlsx", engine= 'xlsxwriter')
weekly_1.to_excel(writer2, sheet_name='Weekly', index=False)
monthly_1.to_excel(writer2, sheet_name='Monthly', index=False)
quarterly_1.to_excel(writer2, sheet_name='Quarterly', index=False)
df_9.to_excel(writer2, sheet_name='YTD', index=False)

writer2.save()


# #sending report to POC cc manager
# data_final=pd.DataFrame([])
# df_5= pd.DataFrame([])
# df_7=pd.DataFrame([])
# df_8=pd.DataFrame([])
# for i in email_details.Manager.unique():
#     data_final=data[(data.Manager==i)]
#     df_5=df5[(df5.Manager==i)]
#     df_7=df7[(df7.Manager==i)]
#     df_8=ytd_final[(ytd2.Manager==i)]
#     writer = pd.ExcelWriter(path1+"/Final_Report"+"/Productivity_Quality_Report_Till_Week"+str(w1)+str(i)+".xlsx", engine='xlsxwriter')
#     data_final.to_excel(writer, sheet_name='Weekly Report', index=False)  # Default position, cell A1.
#     df_5.to_excel(writer, sheet_name='Monthly Report', index=False)
#     df_7.to_excel(writer, sheet_name='Quarterly Report', index=False)
#     df_8.to_excel(writer, sheet_name='YTD', index=False)
#     writer.save()
   
# for i in email_details['POC']:
    
#     receiver_email=email_details.loc[email_details["POC"]==i]["POC_Email"].values[0]
#     cc_email= email_details.loc[email_details["POC"]==i]["Email"].values[0]
#     j= email_details.loc[email_details["POC"]==i]["Manager"].values[0]
#     filename1=(path1+"/Final_Report"+"/Productivity_Quality_Report_Till_Week"+str(w1)+str(j)+".xlsx")
#     email_body= ("PFA the productivity and quality scores till week-"+ str(w1))
    
#     subject= ("Productivity and Quality Report Till.-"+str(w1))
    
#     outlook = win32.Dispatch('outlook.application')
#     mail = outlook.CreateItem(0)
#     mail.To = receiver_email
#     mail.CC = cc_email
#     mail.Subject = str(subject)
#     mail.Body = ("Hello"+ " "+ i  + "," + "\n"+ "\n"+ str(email_body) + "."+ "\n" + "\n"+ "Thank you.")
        

#             # To attach a file to the email (optional):
#     mail.Attachments.Add(filename1)

#     mail.Send()

   
#send email to manager only (both consolidated and detailed one for test run)
weekly_1=pd.DataFrame([])
monthly_1= pd.DataFrame([])
quarterly_1=pd.DataFrame([])
df_9=pd.DataFrame([])
data_final=pd.DataFrame([])
df_5=pd.DataFrame([])
df_7=pd.DataFrame([])
df_8=pd.DataFrame([])
for i in email_details.Manager.unique():
    weekly_1=consolidated_weekly[(consolidated_weekly.Manager==i)]
    monthly_1=monthly_consolidated[(monthly_consolidated.Manager==i)]
    quarterly_1=quarterly_consolidated[(quarterly_consolidated.Manager==i)]
    df_9=consolidated_ytd[(consolidated_ytd.Manager==i)]
    data_final=data[(data.Manager==i)]
    df_5=df5[(df5.Manager==i)]
    df_7=df7[(df7.Manager==i)]
    df_8=ytd_final[(ytd_final.Manager==i)]
    writer = pd.ExcelWriter(path1+"/Final_Report"+"/Productivity_Quality_Report_Till_Week"+str(w1)+str(i)+".xlsx", engine='xlsxwriter')
    data_final.to_excel(writer, sheet_name='Weekly Report', index=False)  # Default position, cell A1.
    df_5.to_excel(writer, sheet_name='Monthly Report', index=False)
    df_7.to_excel(writer, sheet_name='Quarterly Report', index=False)
    df_8.to_excel(writer, sheet_name='YTD', index=False)
    writer.save()
    writer2= pd.ExcelWriter(path1+"/Final_Report"+"/Consolidated_Report_Till_Week_"+str(w1)+str(i)+".xlsx", engine= 'xlsxwriter')
    weekly_1.to_excel(writer2, sheet_name='Weekly', index=False)
    monthly_1.to_excel(writer2, sheet_name='Monthly', index=False)
    quarterly_1.to_excel(writer2, sheet_name='Quarterly', index=False)
    df_9.to_excel(writer2, sheet_name='YTD', index=False)
    writer2.save()
'''    
for i in email_details.Manager.unique():
    
    receiver_email=email_details.loc[email_details["Manager"]==i]["Email"].values[0]
    
    
    filename1=(path1+"/Final_Report"+"/Consolidated_Report_Till_Week_"+str(w1)+str(i)+".xlsx")
    filename2=(path1+"/Final_Report"+"/Productivity_Quality_Report_Till_Week"+str(w1)+str(i)+".xlsx")
    email_body= ("PFA the detailed as well as consolidated productivity and quality scores till week-"+ str(w1)+"."+"\n"+
                 "Please provide POC's login for sending detailed report from next week."
                 "Please feel free to reach out to me in case of any issues/queries.")
    
    subject= ("Productivity and Quality Reports Till.-"+str(w1))
    
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = receiver_email
    
    mail.Subject = str(subject)
    mail.Body = ("Hello"+ " "+ i  + "," + "\n"+ "\n"+ str(email_body) + "."+ "\n" + "\n"+ "Thank you."+"\n"+"Ankita Choudhary")
        

            # To attach a file to the email (optional):
    mail.Attachments.Add(filename1)
    mail.Attachments.Add(filename2)

    mail.Send()
    
# for i in data.Manager.unique():    
#     os.remove(path1+'/Final_Report/'+"Productivity_Quality_Report_Till_Week"+str(w1)+str(i)+".xlsx")

  '''           
    
